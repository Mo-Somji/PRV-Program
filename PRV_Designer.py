import os
os.environ['RPPREFIX'] = r'C:/Program Files (x86)/REFPROP'
from ctREFPROP.ctREFPROP import REFPROPFunctionLibrary
import glob
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from openpyxl import Workbook, load_workbook
from scipy.optimize import fsolve
from decimal import Decimal
wb = load_workbook('Datasheets/Datasheet Template.xlsx')   # loading datasheet template
ws = wb.active

RP = REFPROPFunctionLibrary(os.environ['RPPREFIX'])
RP.SETPATHdll(os.environ['RPPREFIX'])
BASE_UNIT_SI = RP.GETENUMdll(0, 'MASS BASE SI').iEnum
MASS_UNIT_SI = RP.GETENUMdll(0, 'MASS BASE SI').iEnum
MOLAR_UNIT_SI = RP.GETENUMdll(0, 'MOLAR BASE SI').iEnum

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1009, 889)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.gridLayout_10 = QtWidgets.QGridLayout(self.tab)
        self.gridLayout_10.setObjectName("gridLayout_10")
        self.general_info = QtWidgets.QScrollArea(self.tab)
        self.general_info.setWidgetResizable(True)
        self.general_info.setObjectName("general_info")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 948, 955))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout_9 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout_9.setObjectName("gridLayout_9")
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_9.addItem(spacerItem, 5, 0, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_9.addItem(spacerItem1, 11, 0, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_9.addItem(spacerItem2, 14, 0, 1, 1)
        self.connections = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.connections.setObjectName("connections")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.connections)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.inlet_size = QtWidgets.QLabel(self.connections)
        self.inlet_size.setObjectName("inlet_size")
        self.gridLayout_6.addWidget(self.inlet_size, 0, 0, 1, 1)
        self.inlet_size_2 = QtWidgets.QLineEdit(self.connections)
        self.inlet_size_2.setObjectName("inlet_size_2")
        self.gridLayout_6.addWidget(self.inlet_size_2, 0, 1, 1, 1)
        self.outlet_size = QtWidgets.QLabel(self.connections)
        self.outlet_size.setObjectName("outlet_size")
        self.gridLayout_6.addWidget(self.outlet_size, 0, 2, 1, 1)
        self.outlet_size_2 = QtWidgets.QLineEdit(self.connections)
        self.outlet_size_2.setObjectName("outlet_size_2")
        self.gridLayout_6.addWidget(self.outlet_size_2, 0, 3, 1, 1)
        self.inlet_facing = QtWidgets.QLabel(self.connections)
        self.inlet_facing.setObjectName("inlet_facing")
        self.gridLayout_6.addWidget(self.inlet_facing, 1, 0, 1, 1)
        self.inlet_facing_2 = QtWidgets.QComboBox(self.connections)
        self.inlet_facing_2.setObjectName("inlet_facing_2")
        self.inlet_facing_2.addItem("")
        self.inlet_facing_2.addItem("")
        self.inlet_facing_2.addItem("")
        self.inlet_facing_2.addItem("")
        self.inlet_facing_2.addItem("")
        self.inlet_facing_2.addItem("")
        self.gridLayout_6.addWidget(self.inlet_facing_2, 1, 1, 1, 1)
        self.outlet_facing = QtWidgets.QLabel(self.connections)
        self.outlet_facing.setObjectName("outlet_facing")
        self.gridLayout_6.addWidget(self.outlet_facing, 1, 2, 1, 1)
        self.outlet_facing_2 = QtWidgets.QComboBox(self.connections)
        self.outlet_facing_2.setObjectName("outlet_facing_2")
        self.outlet_facing_2.addItem("")
        self.outlet_facing_2.addItem("")
        self.outlet_facing_2.addItem("")
        self.outlet_facing_2.addItem("")
        self.outlet_facing_2.addItem("")
        self.outlet_facing_2.addItem("")
        self.gridLayout_6.addWidget(self.outlet_facing_2, 1, 3, 1, 1)
        self.inlet_facing_3 = QtWidgets.QLineEdit(self.connections)
        self.inlet_facing_3.setEnabled(False)
        self.inlet_facing_3.setReadOnly(True)
        self.inlet_facing_3.setObjectName("inlet_facing_3")
        self.gridLayout_6.addWidget(self.inlet_facing_3, 2, 1, 1, 1)
        self.outlet_facing_3 = QtWidgets.QLineEdit(self.connections)
        self.outlet_facing_3.setEnabled(False)
        self.outlet_facing_3.setReadOnly(True)
        self.outlet_facing_3.setObjectName("outlet_facing_3")
        self.gridLayout_6.addWidget(self.outlet_facing_3, 2, 3, 1, 1)
        self.gridLayout_9.addWidget(self.connections, 10, 0, 1, 1)
        spacerItem3 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_9.addItem(spacerItem3, 7, 0, 1, 1)
        self.selection_basis = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.selection_basis.setObjectName("selection_basis")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.selection_basis)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.asme_code = QtWidgets.QLabel(self.selection_basis)
        self.asme_code.setObjectName("asme_code")
        self.gridLayout_5.addWidget(self.asme_code, 0, 0, 1, 1)
        self.asme_code_2 = QtWidgets.QComboBox(self.selection_basis)
        self.asme_code_2.setMouseTracking(False)
        self.asme_code_2.setFocusPolicy(QtCore.Qt.WheelFocus)
        self.asme_code_2.setAcceptDrops(False)
        self.asme_code_2.setObjectName("asme_code_2")
        self.asme_code_2.addItem("")
        self.asme_code_2.addItem("")
        self.asme_code_2.addItem("")
        self.gridLayout_5.addWidget(self.asme_code_2, 0, 1, 1, 1)
        self.asme_code_3 = QtWidgets.QLineEdit(self.selection_basis)
        self.asme_code_3.setEnabled(False)
        self.asme_code_3.setReadOnly(True)
        self.asme_code_3.setObjectName("asme_code_3")
        self.gridLayout_5.addWidget(self.asme_code_3, 0, 2, 1, 1)
        self.api_code = QtWidgets.QLabel(self.selection_basis)
        self.api_code.setObjectName("api_code")
        self.gridLayout_5.addWidget(self.api_code, 1, 0, 1, 1)
        self.api_code_2 = QtWidgets.QComboBox(self.selection_basis)
        self.api_code_2.setObjectName("api_code_2")
        self.api_code_2.addItem("")
        self.api_code_2.addItem("")
        self.gridLayout_5.addWidget(self.api_code_2, 1, 1, 1, 1)
        self.api_code_3 = QtWidgets.QLineEdit(self.selection_basis)
        self.api_code_3.setEnabled(False)
        self.api_code_3.setReadOnly(True)
        self.api_code_3.setObjectName("api_code_3")
        self.gridLayout_5.addWidget(self.api_code_3, 1, 2, 1, 1)
        self.fire_condition = QtWidgets.QLabel(self.selection_basis)
        self.fire_condition.setObjectName("fire_condition")
        self.gridLayout_5.addWidget(self.fire_condition, 2, 0, 1, 1)
        self.fire_condition_2 = QtWidgets.QComboBox(self.selection_basis)
        self.fire_condition_2.setObjectName("fire_condition_2")
        self.fire_condition_2.addItem("")
        self.fire_condition_2.addItem("")
        self.gridLayout_5.addWidget(self.fire_condition_2, 2, 1, 1, 1)
        self.rupture_disk = QtWidgets.QLabel(self.selection_basis)
        self.rupture_disk.setObjectName("rupture_disk")
        self.gridLayout_5.addWidget(self.rupture_disk, 3, 0, 1, 1)
        self.rupture_disk_2 = QtWidgets.QComboBox(self.selection_basis)
        self.rupture_disk_2.setObjectName("rupture_disk_2")
        self.rupture_disk_2.addItem("")
        self.rupture_disk_2.addItem("")
        self.gridLayout_5.addWidget(self.rupture_disk_2, 3, 1, 1, 1)
        self.gridLayout_9.addWidget(self.selection_basis, 6, 0, 1, 1)
        spacerItem4 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_9.addItem(spacerItem4, 9, 0, 1, 1)
        spacerItem5 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_9.addItem(spacerItem5, 1, 0, 1, 1)
        self.equipment_info = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.equipment_info.setObjectName("equipment_info")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.equipment_info)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.item_number = QtWidgets.QLabel(self.equipment_info)
        self.item_number.setObjectName("item_number")
        self.gridLayout_4.addWidget(self.item_number, 0, 0, 1, 1)
        self.item_number_2 = QtWidgets.QLineEdit(self.equipment_info)
        self.item_number_2.setObjectName("item_number_2")
        self.gridLayout_4.addWidget(self.item_number_2, 0, 1, 1, 1)
        self.service_line_number = QtWidgets.QLabel(self.equipment_info)
        self.service_line_number.setObjectName("service_line_number")
        self.gridLayout_4.addWidget(self.service_line_number, 0, 2, 1, 1)
        self.service_line_number_2 = QtWidgets.QLineEdit(self.equipment_info)
        self.service_line_number_2.setObjectName("service_line_number_2")
        self.gridLayout_4.addWidget(self.service_line_number_2, 0, 3, 1, 1)
        self.tag_number = QtWidgets.QLabel(self.equipment_info)
        self.tag_number.setObjectName("tag_number")
        self.gridLayout_4.addWidget(self.tag_number, 1, 0, 1, 1)
        self.tag_number_2 = QtWidgets.QLineEdit(self.equipment_info)
        self.tag_number_2.setObjectName("tag_number_2")
        self.gridLayout_4.addWidget(self.tag_number_2, 1, 1, 1, 1)
        self.number_required = QtWidgets.QLabel(self.equipment_info)
        self.number_required.setObjectName("number_required")
        self.gridLayout_4.addWidget(self.number_required, 1, 2, 1, 1)
        self.number_required_2 = QtWidgets.QLineEdit(self.equipment_info)
        self.number_required_2.setObjectName("number_required_2")
        self.gridLayout_4.addWidget(self.number_required_2, 1, 3, 1, 1)
        self.gridLayout_9.addWidget(self.equipment_info, 2, 0, 3, 1)
        self.valve_design = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.valve_design.setObjectName("valve_design")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.valve_design)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.seat_tightness = QtWidgets.QLabel(self.valve_design)
        self.seat_tightness.setObjectName("seat_tightness")
        self.gridLayout_2.addWidget(self.seat_tightness, 5, 0, 1, 1)
        self.nozzle_type = QtWidgets.QLabel(self.valve_design)
        self.nozzle_type.setObjectName("nozzle_type")
        self.gridLayout_2.addWidget(self.nozzle_type, 2, 0, 1, 1)
        self.bonnet_type_2 = QtWidgets.QComboBox(self.valve_design)
        self.bonnet_type_2.setObjectName("bonnet_type_2")
        self.bonnet_type_2.addItem("")
        self.bonnet_type_2.addItem("")
        self.gridLayout_2.addWidget(self.bonnet_type_2, 4, 1, 1, 1)
        self.nozzle_type_3 = QtWidgets.QLineEdit(self.valve_design)
        self.nozzle_type_3.setEnabled(False)
        self.nozzle_type_3.setReadOnly(True)
        self.nozzle_type_3.setObjectName("nozzle_type_3")
        self.gridLayout_2.addWidget(self.nozzle_type_3, 3, 1, 1, 1)
        self.seat_tightness_3 = QtWidgets.QLineEdit(self.valve_design)
        self.seat_tightness_3.setEnabled(False)
        self.seat_tightness_3.setReadOnly(True)
        self.seat_tightness_3.setObjectName("seat_tightness_3")
        self.gridLayout_2.addWidget(self.seat_tightness_3, 6, 1, 1, 1)
        self.design_type = QtWidgets.QLabel(self.valve_design)
        self.design_type.setObjectName("design_type")
        self.gridLayout_2.addWidget(self.design_type, 0, 0, 1, 1)
        self.nozzle_type_2 = QtWidgets.QComboBox(self.valve_design)
        self.nozzle_type_2.setObjectName("nozzle_type_2")
        self.nozzle_type_2.addItem("")
        self.nozzle_type_2.addItem("")
        self.nozzle_type_2.addItem("")
        self.gridLayout_2.addWidget(self.nozzle_type_2, 2, 1, 1, 1)
        self.design_type_2 = QtWidgets.QComboBox(self.valve_design)
        self.design_type_2.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.design_type_2.setObjectName("design_type_2")
        self.design_type_2.addItem("")
        self.design_type_2.addItem("")
        self.design_type_2.addItem("")
        self.gridLayout_2.addWidget(self.design_type_2, 0, 1, 1, 1)
        self.bonnet_type = QtWidgets.QLabel(self.valve_design)
        self.bonnet_type.setObjectName("bonnet_type")
        self.gridLayout_2.addWidget(self.bonnet_type, 4, 0, 1, 1)
        self.seat_tightness_2 = QtWidgets.QComboBox(self.valve_design)
        self.seat_tightness_2.setObjectName("seat_tightness_2")
        self.seat_tightness_2.addItem("")
        self.seat_tightness_2.addItem("")
        self.gridLayout_2.addWidget(self.seat_tightness_2, 5, 1, 1, 1)
        self.relief_device_role = QtWidgets.QLabel(self.valve_design)
        self.relief_device_role.setObjectName("relief_device_role")
        self.gridLayout_2.addWidget(self.relief_device_role, 1, 0, 1, 1)
        self.relief_device_role_2 = QtWidgets.QComboBox(self.valve_design)
        self.relief_device_role_2.setObjectName("relief_device_role_2")
        self.relief_device_role_2.addItem("")
        self.relief_device_role_2.addItem("")
        self.relief_device_role_2.addItem("")
        self.gridLayout_2.addWidget(self.relief_device_role_2, 1, 1, 1, 1)
        self.gridLayout_9.addWidget(self.valve_design, 8, 0, 1, 1)
        self.materials = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.materials.setObjectName("materials")
        self.gridLayout_7 = QtWidgets.QGridLayout(self.materials)
        self.gridLayout_7.setObjectName("gridLayout_7")
        self.body = QtWidgets.QLabel(self.materials)
        self.body.setObjectName("body")
        self.gridLayout_7.addWidget(self.body, 0, 0, 1, 1)
        self.body_2 = QtWidgets.QLineEdit(self.materials)
        self.body_2.setText("")
        self.body_2.setObjectName("body_2")
        self.gridLayout_7.addWidget(self.body_2, 0, 1, 1, 1)
        self.bonnet = QtWidgets.QLabel(self.materials)
        self.bonnet.setObjectName("bonnet")
        self.gridLayout_7.addWidget(self.bonnet, 0, 2, 1, 1)
        self.bonnet_2 = QtWidgets.QLineEdit(self.materials)
        self.bonnet_2.setText("")
        self.bonnet_2.setObjectName("bonnet_2")
        self.gridLayout_7.addWidget(self.bonnet_2, 0, 3, 1, 1)
        self.seat = QtWidgets.QLabel(self.materials)
        self.seat.setObjectName("seat")
        self.gridLayout_7.addWidget(self.seat, 0, 4, 1, 1)
        self.seat_2 = QtWidgets.QLineEdit(self.materials)
        self.seat_2.setText("")
        self.seat_2.setObjectName("seat_2")
        self.gridLayout_7.addWidget(self.seat_2, 0, 5, 1, 1)
        self.disk = QtWidgets.QLabel(self.materials)
        self.disk.setObjectName("disk")
        self.gridLayout_7.addWidget(self.disk, 1, 0, 1, 1)
        self.disk_2 = QtWidgets.QLineEdit(self.materials)
        self.disk_2.setText("")
        self.disk_2.setObjectName("disk_2")
        self.gridLayout_7.addWidget(self.disk_2, 1, 1, 1, 1)
        self.resilient_seat = QtWidgets.QLabel(self.materials)
        self.resilient_seat.setObjectName("resilient_seat")
        self.gridLayout_7.addWidget(self.resilient_seat, 1, 2, 1, 1)
        self.resilient_seat_2 = QtWidgets.QLineEdit(self.materials)
        self.resilient_seat_2.setText("")
        self.resilient_seat_2.setObjectName("resilient_seat_2")
        self.gridLayout_7.addWidget(self.resilient_seat_2, 1, 3, 1, 1)
        self.guide = QtWidgets.QLabel(self.materials)
        self.guide.setObjectName("guide")
        self.gridLayout_7.addWidget(self.guide, 1, 4, 1, 1)
        self.guide_2 = QtWidgets.QLineEdit(self.materials)
        self.guide_2.setText("")
        self.guide_2.setObjectName("guide_2")
        self.gridLayout_7.addWidget(self.guide_2, 1, 5, 1, 1)
        self.adjusting_rings = QtWidgets.QLabel(self.materials)
        self.adjusting_rings.setObjectName("adjusting_rings")
        self.gridLayout_7.addWidget(self.adjusting_rings, 2, 0, 1, 1)
        self.adjusting_rings_2 = QtWidgets.QLineEdit(self.materials)
        self.adjusting_rings_2.setText("")
        self.adjusting_rings_2.setObjectName("adjusting_rings_2")
        self.gridLayout_7.addWidget(self.adjusting_rings_2, 2, 1, 1, 1)
        self.spring = QtWidgets.QLabel(self.materials)
        self.spring.setObjectName("spring")
        self.gridLayout_7.addWidget(self.spring, 2, 2, 1, 1)
        self.spring_2 = QtWidgets.QLineEdit(self.materials)
        self.spring_2.setText("")
        self.spring_2.setObjectName("spring_2")
        self.gridLayout_7.addWidget(self.spring_2, 2, 3, 1, 1)
        self.bellows = QtWidgets.QLabel(self.materials)
        self.bellows.setObjectName("bellows")
        self.gridLayout_7.addWidget(self.bellows, 2, 4, 1, 1)
        self.bellows_2 = QtWidgets.QLineEdit(self.materials)
        self.bellows_2.setText("")
        self.bellows_2.setObjectName("bellows_2")
        self.gridLayout_7.addWidget(self.bellows_2, 2, 5, 1, 1)
        self.balanced_piston = QtWidgets.QLabel(self.materials)
        self.balanced_piston.setObjectName("balanced_piston")
        self.gridLayout_7.addWidget(self.balanced_piston, 3, 0, 1, 1)
        self.balanced_piston_2 = QtWidgets.QLineEdit(self.materials)
        self.balanced_piston_2.setText("")
        self.balanced_piston_2.setObjectName("balanced_piston_2")
        self.gridLayout_7.addWidget(self.balanced_piston_2, 3, 1, 1, 1)
        self.nace = QtWidgets.QLabel(self.materials)
        self.nace.setObjectName("nace")
        self.gridLayout_7.addWidget(self.nace, 3, 2, 1, 1)
        self.nace_2 = QtWidgets.QComboBox(self.materials)
        self.nace_2.setObjectName("nace_2")
        self.nace_2.addItem("")
        self.nace_2.addItem("")
        self.gridLayout_7.addWidget(self.nace_2, 3, 3, 1, 1)
        self.internal_gasket = QtWidgets.QLabel(self.materials)
        self.internal_gasket.setObjectName("internal_gasket")
        self.gridLayout_7.addWidget(self.internal_gasket, 3, 4, 1, 1)
        self.internal_gasket_2 = QtWidgets.QLineEdit(self.materials)
        self.internal_gasket_2.setText("")
        self.internal_gasket_2.setObjectName("internal_gasket_2")
        self.gridLayout_7.addWidget(self.internal_gasket_2, 3, 5, 1, 1)
        self.gridLayout_9.addWidget(self.materials, 12, 0, 1, 1)
        self.document_info = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.document_info.setObjectName("document_info")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.document_info)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.sheet_number = QtWidgets.QLabel(self.document_info)
        self.sheet_number.setObjectName("sheet_number")
        self.gridLayout_3.addWidget(self.sheet_number, 0, 0, 1, 1)
        self.sheet_number_2 = QtWidgets.QLineEdit(self.document_info)
        self.sheet_number_2.setObjectName("sheet_number_2")
        self.gridLayout_3.addWidget(self.sheet_number_2, 0, 1, 1, 1)
        self.job_number = QtWidgets.QLabel(self.document_info)
        self.job_number.setObjectName("job_number")
        self.gridLayout_3.addWidget(self.job_number, 0, 2, 1, 1)
        self.job_number_2 = QtWidgets.QLineEdit(self.document_info)
        self.job_number_2.setObjectName("job_number_2")
        self.gridLayout_3.addWidget(self.job_number_2, 0, 3, 1, 1)
        self.requisition_number = QtWidgets.QLabel(self.document_info)
        self.requisition_number.setObjectName("requisition_number")
        self.gridLayout_3.addWidget(self.requisition_number, 1, 0, 1, 1)
        self.requisition_number_2 = QtWidgets.QLineEdit(self.document_info)
        self.requisition_number_2.setObjectName("requisition_number_2")
        self.gridLayout_3.addWidget(self.requisition_number_2, 1, 1, 1, 1)
        self.date = QtWidgets.QLabel(self.document_info)
        self.date.setObjectName("date")
        self.gridLayout_3.addWidget(self.date, 1, 2, 1, 1)
        self.date_2 = QtWidgets.QDateEdit(self.document_info)
        self.date_2.setDateTime(QtCore.QDateTime(QtCore.QDate(2022, 1, 1), QtCore.QTime(0, 0, 0)))
        self.date_2.setObjectName("date_2")
        self.gridLayout_3.addWidget(self.date_2, 1, 3, 1, 1)
        self.revision = QtWidgets.QLabel(self.document_info)
        self.revision.setObjectName("revision")
        self.gridLayout_3.addWidget(self.revision, 2, 0, 1, 1)
        self.revision_2 = QtWidgets.QLineEdit(self.document_info)
        self.revision_2.setObjectName("revision_2")
        self.gridLayout_3.addWidget(self.revision_2, 2, 1, 1, 1)
        self.by = QtWidgets.QLabel(self.document_info)
        self.by.setObjectName("by")
        self.gridLayout_3.addWidget(self.by, 2, 2, 1, 1)
        self.by_2 = QtWidgets.QLineEdit(self.document_info)
        self.by_2.setObjectName("by_2")
        self.gridLayout_3.addWidget(self.by_2, 2, 3, 1, 1)
        self.gridLayout_9.addWidget(self.document_info, 0, 0, 1, 1)
        self.accessories = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.accessories.setObjectName("accessories")
        self.gridLayout_8 = QtWidgets.QGridLayout(self.accessories)
        self.gridLayout_8.setObjectName("gridLayout_8")
        self.cap = QtWidgets.QLabel(self.accessories)
        self.cap.setObjectName("cap")
        self.gridLayout_8.addWidget(self.cap, 0, 0, 1, 1)
        self.cap_2 = QtWidgets.QComboBox(self.accessories)
        self.cap_2.setObjectName("cap_2")
        self.cap_2.addItem("")
        self.cap_2.addItem("")
        self.gridLayout_8.addWidget(self.cap_2, 0, 1, 1, 1)
        self.test_gag = QtWidgets.QLabel(self.accessories)
        self.test_gag.setObjectName("test_gag")
        self.gridLayout_8.addWidget(self.test_gag, 0, 2, 1, 1)
        self.test_gag_2 = QtWidgets.QComboBox(self.accessories)
        self.test_gag_2.setObjectName("test_gag_2")
        self.test_gag_2.addItem("")
        self.test_gag_2.addItem("")
        self.gridLayout_8.addWidget(self.test_gag_2, 0, 3, 1, 1)
        self.lifting_lever = QtWidgets.QLabel(self.accessories)
        self.lifting_lever.setObjectName("lifting_lever")
        self.gridLayout_8.addWidget(self.lifting_lever, 1, 0, 1, 1)
        self.lifting_lever_2 = QtWidgets.QComboBox(self.accessories)
        self.lifting_lever_2.setObjectName("lifting_lever_2")
        self.lifting_lever_2.addItem("")
        self.lifting_lever_2.addItem("")
        self.lifting_lever_2.addItem("")
        self.gridLayout_8.addWidget(self.lifting_lever_2, 1, 1, 1, 1)
        self.bug_screen = QtWidgets.QLabel(self.accessories)
        self.bug_screen.setObjectName("bug_screen")
        self.gridLayout_8.addWidget(self.bug_screen, 1, 2, 1, 1)
        self.bug_screen_2 = QtWidgets.QComboBox(self.accessories)
        self.bug_screen_2.setObjectName("bug_screen_2")
        self.bug_screen_2.addItem("")
        self.bug_screen_2.addItem("")
        self.bug_screen_2.addItem("")
        self.gridLayout_8.addWidget(self.bug_screen_2, 1, 3, 1, 1)
        self.bug_screen_3 = QtWidgets.QLineEdit(self.accessories)
        self.bug_screen_3.setEnabled(True)
        self.bug_screen_3.setReadOnly(False)
        self.bug_screen_3.setObjectName("bug_screen_3")
        self.gridLayout_8.addWidget(self.bug_screen_3, 2, 3, 1, 1)
        self.gridLayout_9.addWidget(self.accessories, 13, 0, 1, 1)
        self.general_info.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout_10.addWidget(self.general_info, 1, 0, 1, 1)
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.gridLayout_11 = QtWidgets.QGridLayout(self.tab_2)
        self.gridLayout_11.setObjectName("gridLayout_11")
        self.fluid_properties = QtWidgets.QScrollArea(self.tab_2)
        self.fluid_properties.setWidgetResizable(True)
        self.fluid_properties.setObjectName("fluid_properties")
        self.scrollAreaWidgetContents_2 = QtWidgets.QWidget()
        self.scrollAreaWidgetContents_2.setGeometry(QtCore.QRect(0, 0, 965, 825))
        self.scrollAreaWidgetContents_2.setObjectName("scrollAreaWidgetContents_2")
        self.unit_basis = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.unit_basis.setGeometry(QtCore.QRect(9, 9, 251, 72))
        self.unit_basis.setObjectName("unit_basis")
        self.gridLayout_12 = QtWidgets.QGridLayout(self.unit_basis)
        self.gridLayout_12.setObjectName("gridLayout_12")
        self.unit_basis_2 = QtWidgets.QLabel(self.unit_basis)
        self.unit_basis_2.setObjectName("unit_basis_2")
        self.gridLayout_12.addWidget(self.unit_basis_2, 0, 0, 1, 1)
        self.unit_basis_3 = QtWidgets.QComboBox(self.unit_basis)
        self.unit_basis_3.setObjectName("unit_basis_3")
        self.unit_basis_3.addItem("")
        self.unit_basis_3.addItem("")
        self.gridLayout_12.addWidget(self.unit_basis_3, 1, 0, 1, 1)
        self.fluid_components = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.fluid_components.setGeometry(QtCore.QRect(9, 87, 951, 251))
        self.fluid_components.setObjectName("fluid_components")
        self.fluid_properties_2 = QtWidgets.QScrollArea(self.fluid_components)
        self.fluid_properties_2.setGeometry(QtCore.QRect(10, 23, 931, 189))
        self.fluid_properties_2.setWidgetResizable(True)
        self.fluid_properties_2.setObjectName("fluid_properties_2")
        self.scrollAreaWidgetContents_3 = QtWidgets.QWidget()
        self.scrollAreaWidgetContents_3.setGeometry(QtCore.QRect(0, 0, 912, 434))
        self.scrollAreaWidgetContents_3.setObjectName("scrollAreaWidgetContents_3")
        self.gridLayout_14 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents_3)
        self.gridLayout_14.setObjectName("gridLayout_14")
        self.fluid_properties_composition = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.fluid_properties_composition.setAlignment(QtCore.Qt.AlignCenter)
        self.fluid_properties_composition.setObjectName("fluid_properties_composition")
        self.gridLayout_14.addWidget(self.fluid_properties_composition, 0, 5, 1, 1)
        self.fluid_properties_name = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.fluid_properties_name.setAlignment(QtCore.Qt.AlignCenter)
        self.fluid_properties_name.setObjectName("fluid_properties_name")
        self.gridLayout_14.addWidget(self.fluid_properties_name, 0, 1, 1, 1)
        self.component_11 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_11.setObjectName("component_11")
        self.gridLayout_14.addWidget(self.component_11, 11, 0, 1, 1)
        self.component_2_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_2_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_2_nist.setObjectName("component_2_nist")
        self.gridLayout_14.addWidget(self.component_2_nist, 2, 3, 1, 1)
        self.component_1_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_1_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_1_nist_identification.setObjectName("component_1_nist_identification")
        self.gridLayout_14.addWidget(self.component_1_nist_identification, 1, 4, 1, 1)
        self.component_4 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_4.setObjectName("component_4")
        self.gridLayout_14.addWidget(self.component_4, 4, 0, 1, 1)
        self.component_8 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_8.setObjectName("component_8")
        self.gridLayout_14.addWidget(self.component_8, 8, 0, 1, 1)
        self.component_12 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_12.setObjectName("component_12")
        self.gridLayout_14.addWidget(self.component_12, 12, 0, 1, 1)
        self.component_1_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_1_name.setObjectName("component_1_name")
        self.gridLayout_14.addWidget(self.component_1_name, 1, 1, 1, 1)
        self.component_1_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_1_composition.setObjectName("component_1_composition")
        self.gridLayout_14.addWidget(self.component_1_composition, 1, 5, 1, 1)
        self.component_13 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_13.setObjectName("component_13")
        self.gridLayout_14.addWidget(self.component_13, 13, 0, 1, 1)
        self.component_2_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_2_name.setObjectName("component_2_name")
        self.gridLayout_14.addWidget(self.component_2_name, 2, 1, 1, 1)
        self.component_10 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_10.setObjectName("component_10")
        self.gridLayout_14.addWidget(self.component_10, 10, 0, 1, 1)
        self.fluid_properties_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.fluid_properties_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.fluid_properties_nist.setObjectName("fluid_properties_nist")
        self.gridLayout_14.addWidget(self.fluid_properties_nist, 0, 3, 1, 1)
        self.component_9 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_9.setObjectName("component_9")
        self.gridLayout_14.addWidget(self.component_9, 9, 0, 1, 1)
        self.component_1_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_1_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_1_nist.setObjectName("component_1_nist")
        self.gridLayout_14.addWidget(self.component_1_nist, 1, 3, 1, 1)
        self.fluid_properties_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.fluid_properties_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.fluid_properties_nist_identification.setObjectName("fluid_properties_nist_identification")
        self.gridLayout_14.addWidget(self.fluid_properties_nist_identification, 0, 4, 1, 1)
        self.component_2_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_2_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_2_nist_identification.setObjectName("component_2_nist_identification")
        self.gridLayout_14.addWidget(self.component_2_nist_identification, 2, 4, 1, 1)
        self.component_7 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_7.setObjectName("component_7")
        self.gridLayout_14.addWidget(self.component_7, 7, 0, 1, 1)
        self.component_5 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_5.setObjectName("component_5")
        self.gridLayout_14.addWidget(self.component_5, 5, 0, 1, 1)
        self.component_6 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_6.setObjectName("component_6")
        self.gridLayout_14.addWidget(self.component_6, 6, 0, 1, 1)
        self.component_2_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_2_composition.setObjectName("component_2_composition")
        self.gridLayout_14.addWidget(self.component_2_composition, 2, 5, 1, 1)
        self.component_1 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_1.setObjectName("component_1")
        self.gridLayout_14.addWidget(self.component_1, 1, 0, 1, 1)
        self.component_14 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_14.setObjectName("component_14")
        self.gridLayout_14.addWidget(self.component_14, 14, 0, 1, 1)
        self.component_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_2.setObjectName("component_2")
        self.gridLayout_14.addWidget(self.component_2, 2, 0, 1, 1)
        self.component_3 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_3.setObjectName("component_3")
        self.gridLayout_14.addWidget(self.component_3, 3, 0, 1, 1)
        self.component_15 = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_15.setObjectName("component_15")
        self.gridLayout_14.addWidget(self.component_15, 15, 0, 1, 1)
        self.component_3_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_3_name.setObjectName("component_3_name")
        self.gridLayout_14.addWidget(self.component_3_name, 3, 1, 1, 1)
        self.component_4_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_4_name.setObjectName("component_4_name")
        self.gridLayout_14.addWidget(self.component_4_name, 4, 1, 1, 1)
        self.component_5_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_5_name.setObjectName("component_5_name")
        self.gridLayout_14.addWidget(self.component_5_name, 5, 1, 1, 1)
        self.component_6_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_6_name.setObjectName("component_6_name")
        self.gridLayout_14.addWidget(self.component_6_name, 6, 1, 1, 1)
        self.component_7_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_7_name.setObjectName("component_7_name")
        self.gridLayout_14.addWidget(self.component_7_name, 7, 1, 1, 1)
        self.component_8_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_8_name.setObjectName("component_8_name")
        self.gridLayout_14.addWidget(self.component_8_name, 8, 1, 1, 1)
        self.component_9_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_9_name.setObjectName("component_9_name")
        self.gridLayout_14.addWidget(self.component_9_name, 9, 1, 1, 1)
        self.component_10_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_10_name.setObjectName("component_10_name")
        self.gridLayout_14.addWidget(self.component_10_name, 10, 1, 1, 1)
        self.component_11_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_11_name.setObjectName("component_11_name")
        self.gridLayout_14.addWidget(self.component_11_name, 11, 1, 1, 1)
        self.component_12_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_12_name.setObjectName("component_12_name")
        self.gridLayout_14.addWidget(self.component_12_name, 12, 1, 1, 1)
        self.component_13_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_13_name.setObjectName("component_13_name")
        self.gridLayout_14.addWidget(self.component_13_name, 13, 1, 1, 1)
        self.component_14_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_14_name.setObjectName("component_14_name")
        self.gridLayout_14.addWidget(self.component_14_name, 14, 1, 1, 1)
        self.component_15_name = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_15_name.setObjectName("component_15_name")
        self.gridLayout_14.addWidget(self.component_15_name, 15, 1, 1, 1)
        self.component_3_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_3_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_3_nist.setObjectName("component_3_nist")
        self.gridLayout_14.addWidget(self.component_3_nist, 3, 3, 1, 1)
        self.component_4_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_4_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_4_nist.setObjectName("component_4_nist")
        self.gridLayout_14.addWidget(self.component_4_nist, 4, 3, 1, 1)
        self.component_5_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_5_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_5_nist.setObjectName("component_5_nist")
        self.gridLayout_14.addWidget(self.component_5_nist, 5, 3, 1, 1)
        self.component_6_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_6_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_6_nist.setObjectName("component_6_nist")
        self.gridLayout_14.addWidget(self.component_6_nist, 6, 3, 1, 1)
        self.component_7_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_7_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_7_nist.setObjectName("component_7_nist")
        self.gridLayout_14.addWidget(self.component_7_nist, 7, 3, 1, 1)
        self.component_8_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_8_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_8_nist.setObjectName("component_8_nist")
        self.gridLayout_14.addWidget(self.component_8_nist, 8, 3, 1, 1)
        self.component_9_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_9_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_9_nist.setObjectName("component_9_nist")
        self.gridLayout_14.addWidget(self.component_9_nist, 9, 3, 1, 1)
        self.component_10_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_10_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_10_nist.setObjectName("component_10_nist")
        self.gridLayout_14.addWidget(self.component_10_nist, 10, 3, 1, 1)
        self.component_11_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_11_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_11_nist.setObjectName("component_11_nist")
        self.gridLayout_14.addWidget(self.component_11_nist, 11, 3, 1, 1)
        self.component_12_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_12_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_12_nist.setObjectName("component_12_nist")
        self.gridLayout_14.addWidget(self.component_12_nist, 12, 3, 1, 1)
        self.component_13_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_13_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_13_nist.setObjectName("component_13_nist")
        self.gridLayout_14.addWidget(self.component_13_nist, 13, 3, 1, 1)
        self.component_14_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_14_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_14_nist.setObjectName("component_14_nist")
        self.gridLayout_14.addWidget(self.component_14_nist, 14, 3, 1, 1)
        self.component_15_nist = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_15_nist.setAlignment(QtCore.Qt.AlignCenter)
        self.component_15_nist.setObjectName("component_15_nist")
        self.gridLayout_14.addWidget(self.component_15_nist, 15, 3, 1, 1)
        self.component_3_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_3_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_3_nist_identification.setObjectName("component_3_nist_identification")
        self.gridLayout_14.addWidget(self.component_3_nist_identification, 3, 4, 1, 1)
        self.component_4_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_4_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_4_nist_identification.setObjectName("component_4_nist_identification")
        self.gridLayout_14.addWidget(self.component_4_nist_identification, 4, 4, 1, 1)
        self.component_5_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_5_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_5_nist_identification.setObjectName("component_5_nist_identification")
        self.gridLayout_14.addWidget(self.component_5_nist_identification, 5, 4, 1, 1)
        self.component_6_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_6_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_6_nist_identification.setObjectName("component_6_nist_identification")
        self.gridLayout_14.addWidget(self.component_6_nist_identification, 6, 4, 1, 1)
        self.component_7_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_7_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_7_nist_identification.setObjectName("component_7_nist_identification")
        self.gridLayout_14.addWidget(self.component_7_nist_identification, 7, 4, 1, 1)
        self.component_8_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_8_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_8_nist_identification.setObjectName("component_8_nist_identification")
        self.gridLayout_14.addWidget(self.component_8_nist_identification, 8, 4, 1, 1)
        self.component_9_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_9_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_9_nist_identification.setObjectName("component_9_nist_identification")
        self.gridLayout_14.addWidget(self.component_9_nist_identification, 9, 4, 1, 1)
        self.component_10_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_10_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_10_nist_identification.setObjectName("component_10_nist_identification")
        self.gridLayout_14.addWidget(self.component_10_nist_identification, 10, 4, 1, 1)
        self.component_11_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_11_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_11_nist_identification.setObjectName("component_11_nist_identification")
        self.gridLayout_14.addWidget(self.component_11_nist_identification, 11, 4, 1, 1)
        self.component_12_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_12_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_12_nist_identification.setObjectName("component_12_nist_identification")
        self.gridLayout_14.addWidget(self.component_12_nist_identification, 12, 4, 1, 1)
        self.component_13_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_13_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_13_nist_identification.setObjectName("component_13_nist_identification")
        self.gridLayout_14.addWidget(self.component_13_nist_identification, 13, 4, 1, 1)
        self.component_14_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_14_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_14_nist_identification.setObjectName("component_14_nist_identification")
        self.gridLayout_14.addWidget(self.component_14_nist_identification, 14, 4, 1, 1)
        self.component_15_nist_identification = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.component_15_nist_identification.setAlignment(QtCore.Qt.AlignCenter)
        self.component_15_nist_identification.setObjectName("component_15_nist_identification")
        self.gridLayout_14.addWidget(self.component_15_nist_identification, 15, 4, 1, 1)
        self.component_3_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_3_composition.setObjectName("component_3_composition")
        self.gridLayout_14.addWidget(self.component_3_composition, 3, 5, 1, 1)
        self.component_4_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_4_composition.setObjectName("component_4_composition")
        self.gridLayout_14.addWidget(self.component_4_composition, 4, 5, 1, 1)
        self.component_5_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_5_composition.setObjectName("component_5_composition")
        self.gridLayout_14.addWidget(self.component_5_composition, 5, 5, 1, 1)
        self.component_6_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_6_composition.setObjectName("component_6_composition")
        self.gridLayout_14.addWidget(self.component_6_composition, 6, 5, 1, 1)
        self.component_7_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_7_composition.setObjectName("component_7_composition")
        self.gridLayout_14.addWidget(self.component_7_composition, 7, 5, 1, 1)
        self.component_8_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_8_composition.setObjectName("component_8_composition")
        self.gridLayout_14.addWidget(self.component_8_composition, 8, 5, 1, 1)
        self.component_9_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_9_composition.setObjectName("component_9_composition")
        self.gridLayout_14.addWidget(self.component_9_composition, 9, 5, 1, 1)
        self.component_10_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_10_composition.setObjectName("component_10_composition")
        self.gridLayout_14.addWidget(self.component_10_composition, 10, 5, 1, 1)
        self.component_11_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_11_composition.setObjectName("component_11_composition")
        self.gridLayout_14.addWidget(self.component_11_composition, 11, 5, 1, 1)
        self.component_12_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_12_composition.setObjectName("component_12_composition")
        self.gridLayout_14.addWidget(self.component_12_composition, 12, 5, 1, 1)
        self.component_13_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_13_composition.setObjectName("component_13_composition")
        self.gridLayout_14.addWidget(self.component_13_composition, 13, 5, 1, 1)
        self.component_14_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_14_composition.setObjectName("component_14_composition")
        self.gridLayout_14.addWidget(self.component_14_composition, 14, 5, 1, 1)
        self.component_15_composition = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.component_15_composition.setObjectName("component_15_composition")
        self.gridLayout_14.addWidget(self.component_15_composition, 15, 5, 1, 1)
        self.fluid_properties_2.setWidget(self.scrollAreaWidgetContents_3)
        self.total_composition = QtWidgets.QLabel(self.fluid_components)
        self.total_composition.setGeometry(QtCore.QRect(580, 220, 111, 20))
        self.total_composition.setObjectName("total_composition")
        self.total_composition_2 = QtWidgets.QLineEdit(self.fluid_components)
        self.total_composition_2.setEnabled(False)
        self.total_composition_2.setGeometry(QtCore.QRect(683, 220, 231, 20))
        self.total_composition_2.setObjectName("total_composition_2")
        self.conditions = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.conditions.setGeometry(QtCore.QRect(10, 350, 601, 191))
        self.conditions.setObjectName("conditions")
        self.gridLayout_15 = QtWidgets.QGridLayout(self.conditions)
        self.gridLayout_15.setObjectName("gridLayout_15")
        self.total_back_pressure = QtWidgets.QLabel(self.conditions)
        self.total_back_pressure.setObjectName("total_back_pressure")
        self.gridLayout_15.addWidget(self.total_back_pressure, 4, 0, 1, 1)
        self.relieving_pressure_2 = QtWidgets.QLineEdit(self.conditions)
        self.relieving_pressure_2.setEnabled(True)
        self.relieving_pressure_2.setText("")
        self.relieving_pressure_2.setObjectName("relieving_pressure_2")
        self.gridLayout_15.addWidget(self.relieving_pressure_2, 5, 1, 1, 1)
        self.set_pressure_2 = QtWidgets.QLineEdit(self.conditions)
        self.set_pressure_2.setText("")
        self.set_pressure_2.setObjectName("set_pressure_2")
        self.gridLayout_15.addWidget(self.set_pressure_2, 2, 1, 1, 1)
        self.relieving_temp_2 = QtWidgets.QLineEdit(self.conditions)
        self.relieving_temp_2.setObjectName("relieving_temp_2")
        self.gridLayout_15.addWidget(self.relieving_temp_2, 0, 1, 1, 1)
        self.overpressure_percentage_2 = QtWidgets.QLineEdit(self.conditions)
        self.overpressure_percentage_2.setText("")
        self.overpressure_percentage_2.setObjectName("overpressure_percentage_2")
        self.gridLayout_15.addWidget(self.overpressure_percentage_2, 3, 1, 1, 1)
        self.total_back_pressure_2 = QtWidgets.QLineEdit(self.conditions)
        self.total_back_pressure_2.setObjectName("total_back_pressure_2")
        self.gridLayout_15.addWidget(self.total_back_pressure_2, 4, 1, 1, 1)
        self.set_pressure = QtWidgets.QLabel(self.conditions)
        self.set_pressure.setObjectName("set_pressure")
        self.gridLayout_15.addWidget(self.set_pressure, 2, 0, 1, 1)
        self.relieving_temp = QtWidgets.QLabel(self.conditions)
        self.relieving_temp.setObjectName("relieving_temp")
        self.gridLayout_15.addWidget(self.relieving_temp, 0, 0, 1, 1)
        self.relieving_pressure = QtWidgets.QLabel(self.conditions)
        self.relieving_pressure.setObjectName("relieving_pressure")
        self.gridLayout_15.addWidget(self.relieving_pressure, 5, 0, 1, 1)
        self.overpressure_percentage = QtWidgets.QLabel(self.conditions)
        self.overpressure_percentage.setObjectName("overpressure_percentage")
        self.gridLayout_15.addWidget(self.overpressure_percentage, 3, 0, 1, 1)
        self.equipment_mawp = QtWidgets.QLabel(self.conditions)
        self.equipment_mawp.setObjectName("equipment_mawp")
        self.gridLayout_15.addWidget(self.equipment_mawp, 1, 0, 1, 1)
        self.equipment_mawp_2 = QtWidgets.QLineEdit(self.conditions)
        self.equipment_mawp_2.setObjectName("equipment_mawp_2")
        self.gridLayout_15.addWidget(self.equipment_mawp_2, 1, 1, 1, 1)
        self.fluid_properties_results = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.fluid_properties_results.setGeometry(QtCore.QRect(10, 600, 511, 211))
        self.fluid_properties_results.setObjectName("fluid_properties_results")
        self.gridLayout_16 = QtWidgets.QGridLayout(self.fluid_properties_results)
        self.gridLayout_16.setObjectName("gridLayout_16")
        self.molecular_weight_2 = QtWidgets.QLineEdit(self.fluid_properties_results)
        self.molecular_weight_2.setObjectName("molecular_weight_2")
        self.gridLayout_16.addWidget(self.molecular_weight_2, 4, 1, 1, 1)
        self.density_2 = QtWidgets.QLineEdit(self.fluid_properties_results)
        self.density_2.setObjectName("density_2")
        self.gridLayout_16.addWidget(self.density_2, 5, 1, 1, 1)
        self.specific_heat_cp_2 = QtWidgets.QLineEdit(self.fluid_properties_results)
        self.specific_heat_cp_2.setObjectName("specific_heat_cp_2")
        self.gridLayout_16.addWidget(self.specific_heat_cp_2, 1, 1, 1, 1)
        self.density = QtWidgets.QLabel(self.fluid_properties_results)
        self.density.setObjectName("density")
        self.gridLayout_16.addWidget(self.density, 5, 0, 1, 1)
        self.compressibility_factor = QtWidgets.QLabel(self.fluid_properties_results)
        self.compressibility_factor.setObjectName("compressibility_factor")
        self.gridLayout_16.addWidget(self.compressibility_factor, 3, 0, 1, 1)
        self.specific_heat_cp = QtWidgets.QLabel(self.fluid_properties_results)
        self.specific_heat_cp.setObjectName("specific_heat_cp")
        self.gridLayout_16.addWidget(self.specific_heat_cp, 1, 0, 1, 1)
        self.compressibility_factor_2 = QtWidgets.QLineEdit(self.fluid_properties_results)
        self.compressibility_factor_2.setObjectName("compressibility_factor_2")
        self.gridLayout_16.addWidget(self.compressibility_factor_2, 3, 1, 1, 1)
        self.fluid_phase_2 = QtWidgets.QComboBox(self.fluid_properties_results)
        self.fluid_phase_2.setObjectName("fluid_phase_2")
        self.fluid_phase_2.addItem("")
        self.fluid_phase_2.addItem("")
        self.fluid_phase_2.addItem("")
        self.fluid_phase_2.addItem("")
        self.fluid_phase_2.addItem("")
        self.gridLayout_16.addWidget(self.fluid_phase_2, 0, 1, 1, 1)
        self.fluid_phase = QtWidgets.QLabel(self.fluid_properties_results)
        self.fluid_phase.setObjectName("fluid_phase")
        self.gridLayout_16.addWidget(self.fluid_phase, 0, 0, 1, 1)
        self.molecular_weight = QtWidgets.QLabel(self.fluid_properties_results)
        self.molecular_weight.setObjectName("molecular_weight")
        self.gridLayout_16.addWidget(self.molecular_weight, 4, 0, 1, 1)
        self.specific_heat_cv = QtWidgets.QLabel(self.fluid_properties_results)
        self.specific_heat_cv.setObjectName("specific_heat_cv")
        self.gridLayout_16.addWidget(self.specific_heat_cv, 2, 0, 1, 1)
        self.specific_heat_cv_2 = QtWidgets.QLineEdit(self.fluid_properties_results)
        self.specific_heat_cv_2.setObjectName("specific_heat_cv_2")
        self.gridLayout_16.addWidget(self.specific_heat_cv_2, 2, 1, 1, 1)
        self.viscosity = QtWidgets.QLabel(self.fluid_properties_results)
        self.viscosity.setObjectName("viscosity")
        self.gridLayout_16.addWidget(self.viscosity, 6, 0, 1, 1)
        self.viscosity_2 = QtWidgets.QLineEdit(self.fluid_properties_results)
        self.viscosity_2.setObjectName("viscosity_2")
        self.gridLayout_16.addWidget(self.viscosity_2, 6, 1, 1, 1)
        self.nist_properties_generation = QtWidgets.QPushButton(self.scrollAreaWidgetContents_2)
        self.nist_properties_generation.setEnabled(False)
        self.nist_properties_generation.setGeometry(QtCore.QRect(10, 560, 331, 23))
        self.nist_properties_generation.setObjectName("nist_properties_generation")
        self.fluid_properties.setWidget(self.scrollAreaWidgetContents_2)
        self.gridLayout_11.addWidget(self.fluid_properties, 0, 0, 1, 1)
        self.tabWidget.addTab(self.tab_2, "")
        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.generate_datasheet = QtWidgets.QPushButton(self.tab_3)
        self.generate_datasheet.setGeometry(QtCore.QRect(670, 430, 121, 31))
        self.generate_datasheet.setObjectName("generate_datasheet")
        self.variables = QtWidgets.QGroupBox(self.tab_3)
        self.variables.setGeometry(QtCore.QRect(20, 510, 461, 171))
        self.variables.setObjectName("variables")
        self.gridLayout_17 = QtWidgets.QGridLayout(self.variables)
        self.gridLayout_17.setObjectName("gridLayout_17")
        self.Kc = QtWidgets.QLabel(self.variables)
        self.Kc.setObjectName("Kc")
        self.gridLayout_17.addWidget(self.Kc, 2, 0, 1, 1)
        self.Kb = QtWidgets.QLabel(self.variables)
        self.Kb.setObjectName("Kb")
        self.gridLayout_17.addWidget(self.Kb, 1, 0, 1, 1)
        self.mass_flow = QtWidgets.QLabel(self.variables)
        self.mass_flow.setObjectName("mass_flow")
        self.gridLayout_17.addWidget(self.mass_flow, 3, 0, 1, 1)
        self.mass_flow_2 = QtWidgets.QLineEdit(self.variables)
        self.mass_flow_2.setObjectName("mass_flow_2")
        self.gridLayout_17.addWidget(self.mass_flow_2, 3, 1, 1, 1)
        self.Kb_2 = QtWidgets.QLineEdit(self.variables)
        self.Kb_2.setObjectName("Kb_2")
        self.gridLayout_17.addWidget(self.Kb_2, 1, 1, 1, 1)
        self.Kd_2 = QtWidgets.QLineEdit(self.variables)
        self.Kd_2.setObjectName("Kd_2")
        self.gridLayout_17.addWidget(self.Kd_2, 0, 1, 1, 1)
        self.Kc_2 = QtWidgets.QLineEdit(self.variables)
        self.Kc_2.setObjectName("Kc_2")
        self.gridLayout_17.addWidget(self.Kc_2, 2, 1, 1, 1)
        self.Kd = QtWidgets.QLabel(self.variables)
        self.Kd.setObjectName("Kd")
        self.gridLayout_17.addWidget(self.Kd, 0, 0, 1, 1)
        self.vol_flow = QtWidgets.QLabel(self.variables)
        self.vol_flow.setObjectName("vol_flow")
        self.gridLayout_17.addWidget(self.vol_flow, 4, 0, 1, 1)
        self.vol_flow_2 = QtWidgets.QLineEdit(self.variables)
        self.vol_flow_2.setEnabled(False)
        self.vol_flow_2.setObjectName("vol_flow_2")
        self.gridLayout_17.addWidget(self.vol_flow_2, 4, 1, 1, 1)
        self.steam_conditions = QtWidgets.QGroupBox(self.tab_3)
        self.steam_conditions.setGeometry(QtCore.QRect(20, 210, 461, 91))
        self.steam_conditions.setObjectName("steam_conditions")
        self.gridLayout_19 = QtWidgets.QGridLayout(self.steam_conditions)
        self.gridLayout_19.setObjectName("gridLayout_19")
        self.KN = QtWidgets.QLabel(self.steam_conditions)
        self.KN.setObjectName("KN")
        self.gridLayout_19.addWidget(self.KN, 0, 0, 1, 1)
        self.KN_2 = QtWidgets.QLineEdit(self.steam_conditions)
        self.KN_2.setEnabled(False)
        self.KN_2.setObjectName("KN_2")
        self.gridLayout_19.addWidget(self.KN_2, 0, 1, 1, 1)
        self.KSH = QtWidgets.QLabel(self.steam_conditions)
        self.KSH.setObjectName("KSH")
        self.gridLayout_19.addWidget(self.KSH, 1, 0, 1, 1)
        self.KSH_2 = QtWidgets.QLineEdit(self.steam_conditions)
        self.KSH_2.setEnabled(False)
        self.KSH_2.setObjectName("KSH_2")
        self.gridLayout_19.addWidget(self.KSH_2, 1, 1, 1, 1)
        self.liquid_phase_conditions = QtWidgets.QGroupBox(self.tab_3)
        self.liquid_phase_conditions.setGeometry(QtCore.QRect(20, 320, 461, 181))
        self.liquid_phase_conditions.setObjectName("liquid_phase_conditions")
        self.gridLayout_18 = QtWidgets.QGridLayout(self.liquid_phase_conditions)
        self.gridLayout_18.setObjectName("gridLayout_18")
        self.Kp = QtWidgets.QLabel(self.liquid_phase_conditions)
        self.Kp.setObjectName("Kp")
        self.gridLayout_18.addWidget(self.Kp, 1, 0, 1, 1)
        self.capacity_certification_requirement_2 = QtWidgets.QComboBox(self.liquid_phase_conditions)
        self.capacity_certification_requirement_2.setEnabled(False)
        self.capacity_certification_requirement_2.setObjectName("capacity_certification_requirement_2")
        self.capacity_certification_requirement_2.addItem("")
        self.capacity_certification_requirement_2.addItem("")
        self.gridLayout_18.addWidget(self.capacity_certification_requirement_2, 0, 1, 1, 1)
        self.Kv_2 = QtWidgets.QLineEdit(self.liquid_phase_conditions)
        self.Kv_2.setEnabled(False)
        self.Kv_2.setObjectName("Kv_2")
        self.gridLayout_18.addWidget(self.Kv_2, 4, 1, 1, 1)
        self.capacity_certification_requirement = QtWidgets.QLabel(self.liquid_phase_conditions)
        self.capacity_certification_requirement.setObjectName("capacity_certification_requirement")
        self.gridLayout_18.addWidget(self.capacity_certification_requirement, 0, 0, 1, 1)
        self.Kw = QtWidgets.QLabel(self.liquid_phase_conditions)
        self.Kw.setObjectName("Kw")
        self.gridLayout_18.addWidget(self.Kw, 2, 0, 1, 1)
        self.Kw_2 = QtWidgets.QLineEdit(self.liquid_phase_conditions)
        self.Kw_2.setEnabled(False)
        self.Kw_2.setObjectName("Kw_2")
        self.gridLayout_18.addWidget(self.Kw_2, 2, 1, 1, 1)
        self.Kp_2 = QtWidgets.QLineEdit(self.liquid_phase_conditions)
        self.Kp_2.setEnabled(False)
        self.Kp_2.setObjectName("Kp_2")
        self.gridLayout_18.addWidget(self.Kp_2, 1, 1, 1, 1)
        self.Kv = QtWidgets.QLabel(self.liquid_phase_conditions)
        self.Kv.setObjectName("Kv")
        self.gridLayout_18.addWidget(self.Kv, 4, 0, 1, 1)
        self.Re = QtWidgets.QLabel(self.liquid_phase_conditions)
        self.Re.setObjectName("Re")
        self.gridLayout_18.addWidget(self.Re, 3, 0, 1, 1)
        self.Re_2 = QtWidgets.QLineEdit(self.liquid_phase_conditions)
        self.Re_2.setEnabled(False)
        self.Re_2.setObjectName("Re_2")
        self.gridLayout_18.addWidget(self.Re_2, 3, 1, 1, 1)
        self.vapour_conditions = QtWidgets.QGroupBox(self.tab_3)
        self.vapour_conditions.setGeometry(QtCore.QRect(20, 20, 461, 181))
        self.vapour_conditions.setObjectName("vapour_conditions")
        self.gridLayout_13 = QtWidgets.QGridLayout(self.vapour_conditions)
        self.gridLayout_13.setObjectName("gridLayout_13")
        self.pcf = QtWidgets.QLabel(self.vapour_conditions)
        self.pcf.setObjectName("pcf")
        self.gridLayout_13.addWidget(self.pcf, 1, 0, 1, 1)
        self.critical_flow_coefficient = QtWidgets.QLabel(self.vapour_conditions)
        self.critical_flow_coefficient.setObjectName("critical_flow_coefficient")
        self.gridLayout_13.addWidget(self.critical_flow_coefficient, 3, 0, 1, 1)
        self.k_2 = QtWidgets.QLineEdit(self.vapour_conditions)
        self.k_2.setEnabled(False)
        self.k_2.setObjectName("k_2")
        self.gridLayout_13.addWidget(self.k_2, 0, 1, 1, 1)
        self.critical_flow_coefficient_2 = QtWidgets.QLineEdit(self.vapour_conditions)
        self.critical_flow_coefficient_2.setEnabled(False)
        self.critical_flow_coefficient_2.setObjectName("critical_flow_coefficient_2")
        self.gridLayout_13.addWidget(self.critical_flow_coefficient_2, 3, 1, 1, 1)
        self.vapour_state = QtWidgets.QLabel(self.vapour_conditions)
        self.vapour_state.setObjectName("vapour_state")
        self.gridLayout_13.addWidget(self.vapour_state, 2, 0, 1, 1)
        self.k = QtWidgets.QLabel(self.vapour_conditions)
        self.k.setObjectName("k")
        self.gridLayout_13.addWidget(self.k, 0, 0, 1, 1)
        self.pcf_2 = QtWidgets.QLineEdit(self.vapour_conditions)
        self.pcf_2.setEnabled(False)
        self.pcf_2.setObjectName("pcf_2")
        self.gridLayout_13.addWidget(self.pcf_2, 1, 1, 1, 1)
        self.vapour_state_2 = QtWidgets.QComboBox(self.vapour_conditions)
        self.vapour_state_2.setEnabled(True)
        self.vapour_state_2.setObjectName("vapour_state_2")
        self.vapour_state_2.addItem("")
        self.vapour_state_2.addItem("")
        self.vapour_state_2.addItem("")
        self.gridLayout_13.addWidget(self.vapour_state_2, 2, 1, 1, 1)
        self.subcritical_flow_coefficient = QtWidgets.QLabel(self.vapour_conditions)
        self.subcritical_flow_coefficient.setObjectName("subcritical_flow_coefficient")
        self.gridLayout_13.addWidget(self.subcritical_flow_coefficient, 4, 0, 1, 1)
        self.subcritical_flow_coefficient_2 = QtWidgets.QLineEdit(self.vapour_conditions)
        self.subcritical_flow_coefficient_2.setEnabled(False)
        self.subcritical_flow_coefficient_2.setObjectName("subcritical_flow_coefficient_2")
        self.gridLayout_13.addWidget(self.subcritical_flow_coefficient_2, 4, 1, 1, 1)
        self.warning_error_messages = QtWidgets.QGroupBox(self.tab_3)
        self.warning_error_messages.setGeometry(QtCore.QRect(540, 30, 381, 291))
        self.warning_error_messages.setObjectName("warning_error_messages")
        self.results = QtWidgets.QGroupBox(self.tab_3)
        self.results.setGeometry(QtCore.QRect(540, 340, 381, 71))
        self.results.setObjectName("results")
        self.gridLayout_20 = QtWidgets.QGridLayout(self.results)
        self.gridLayout_20.setObjectName("gridLayout_20")
        self.discharge_area = QtWidgets.QLabel(self.results)
        self.discharge_area.setObjectName("discharge_area")
        self.gridLayout_20.addWidget(self.discharge_area, 0, 0, 1, 1)
        self.discharge_area_2 = QtWidgets.QLineEdit(self.results)
        self.discharge_area_2.setEnabled(False)
        self.discharge_area_2.setObjectName("discharge_area_2")
        self.gridLayout_20.addWidget(self.discharge_area_2, 0, 1, 1, 1)
        self.tabWidget.addTab(self.tab_3, "")
        self.gridLayout.addWidget(self.tabWidget, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

# Start of editing UI Code ************************************************************************************************
        self.asme_code_2.activated[str].connect(self.Asme_Code_UI_Logic)            # ASME Combobox Logic Func
        self.api_code_2.activated[str].connect(self.API_Code_UI_Logic)              # API 526 Combobox Logic Func
        self.nozzle_type_2.activated[str].connect(self.Nozzle_Type_Logic)           # nozzle type combobox logic func
        self.seat_tightness_2.activated[str].connect(self.Seat_Tightness)           # seat tightness combobox logic func
        self.inlet_facing_2.activated[str].connect(self.Inlet_Facing)               # inlet flange rating combobox logic func
        self.outlet_facing_2.activated[str].connect(self.Outlet_Facing)             # outlet flange rating combobox logic func
        self.bug_screen_2.activated[str].connect(self.Bug_Screen)                   # bug screen combobox logic func
# End of editing UI Code **************************************************************************************************

# Start of fluid components and compositions section **********************************************************************
        # Function called for mass or molar units
        self.unit_basis_3.activated[str].connect(self.Unit_Basis)

        # Function called for component name NIST check
        self.component_1_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_2_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_3_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_4_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_5_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_6_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_7_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_8_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_9_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_10_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_11_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_12_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_13_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_14_name.editingFinished.connect(self.Component_NIST_Check)
        self.component_15_name.editingFinished.connect(self.Component_NIST_Check)

        # Function called for component name repeat check
        self.component_1_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_2_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_3_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_4_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_5_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_6_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_7_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_8_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_9_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_10_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_11_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_12_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_13_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_14_name.editingFinished.connect(self.Component_Name_Repeat_Check)
        self.component_15_name.editingFinished.connect(self.Component_Name_Repeat_Check)

        # Function called for composition number test
        self.component_1_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_2_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_3_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_4_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_5_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_6_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_7_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_8_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_9_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_10_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_11_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_12_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_13_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_14_composition.editingFinished.connect(self.Composition_Number_Test)
        self.component_15_composition.editingFinished.connect(self.Composition_Number_Test)

        # Function called for composition sum test
        self.component_1_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_2_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_3_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_4_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_5_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_6_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_7_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_8_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_9_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_10_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_11_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_12_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_13_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_14_composition.editingFinished.connect(self.Composition_Sum_Test)
        self.component_15_composition.editingFinished.connect(self.Composition_Sum_Test)

        # Function called for activation of 'generate NIST properties' button
        self.component_1_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_2_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_3_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_4_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_5_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_6_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_7_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_8_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_9_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_10_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_11_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_12_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_13_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_14_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_15_name.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_1_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_2_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_3_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_4_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_5_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_6_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_7_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_8_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_9_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_10_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_11_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_12_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_13_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_14_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.component_15_composition.editingFinished.connect(self.NIST_Properties_Generation)
        self.relieving_temp_2.editingFinished.connect(self.NIST_Properties_Generation)
        self.equipment_mawp_2.editingFinished.connect(self.NIST_Properties_Generation)
        self.set_pressure_2.editingFinished.connect(self.NIST_Properties_Generation)
        self.overpressure_percentage_2.editingFinished.connect(self.NIST_Properties_Generation)
        self.total_back_pressure_2.editingFinished.connect(self.NIST_Properties_Generation)
        #self.relieving_pressure_2.editingFinished.connect(self.NIST_Properties_Generation)

        # Function called for activation of 'conditions' number check
        self.relieving_temp_2.editingFinished.connect(self.Conditions_Number_Check)
        self.equipment_mawp_2.editingFinished.connect(self.Conditions_Number_Check)
        self.set_pressure_2.editingFinished.connect(self.Conditions_Number_Check)
        self.overpressure_percentage_2.editingFinished.connect(self.Conditions_Number_Check)
        self.total_back_pressure_2.editingFinished.connect(self.Conditions_Number_Check)

        # Function called for when equipment MAWP is inputted'
        #self.tabWidget.currentChanged.connect(self.Equipment_MAWP)
        self.relief_device_role_2.activated[str].connect(self.Equipment_MAWP)
        self.equipment_mawp_2.editingFinished.connect(self.Equipment_MAWP)

        # Function called for when PRV set pressure is inputted manually
        self.equipment_mawp_2.editingFinished.connect(self.Set_Pressure)
        self.set_pressure_2.editingFinished.connect(self.Set_Pressure)

        # Function called for calculating 'maximum allowable accumulated pressure %'
        self.tabWidget.currentChanged.connect(self.Overpressure_Percentage)
        self.relieving_temp_2.editingFinished.connect(self.Overpressure_Percentage)
        self.equipment_mawp_2.editingFinished.connect(self.Overpressure_Percentage)
        self.set_pressure_2.editingFinished.connect(self.Overpressure_Percentage)
        self.overpressure_percentage_2.editingFinished.connect(self.Overpressure_Percentage)
        self.total_back_pressure_2.editingFinished.connect(self.Overpressure_Percentage)

        # Function called for calculating initial value for 'relieving pressure P1'
        #self.tabWidget.currentChanged.connect(self.Relieving_Pressure_Calc)
        self.relieving_temp_2.editingFinished.connect(self.Relieving_Pressure_Calc)
        self.equipment_mawp_2.editingFinished.connect(self.Relieving_Pressure_Calc)
        self.set_pressure_2.editingFinished.connect(self.Relieving_Pressure_Calc)
        self.overpressure_percentage_2.editingFinished.connect(self.Relieving_Pressure_Calc)
        self.total_back_pressure_2.editingFinished.connect(self.Relieving_Pressure_Calc)

        # Function called for calculating value for 'relieving pressure P1'
        self.relieving_pressure_2.editingFinished.connect(self.Relieving_Pressure_Calc_2)

        # Function called for pressing the 'generate properties from nist refprop' button
        self.nist_properties_generation.clicked.connect(self.NIST_Properties)

        # Function called for editing fluid properties manually if chosen
        self.total_back_pressure_2.editingFinished.connect(self.Manual_Properties)
        self.relieving_pressure_2.editingFinished.connect(self.Manual_Properties)
        self.specific_heat_cp_2.editingFinished.connect(self.Manual_Properties)
        self.specific_heat_cv_2.editingFinished.connect(self.Manual_Properties)
        self.fluid_phase_2.activated[str].connect(self.Manual_Properties)
        self.specific_heat_cp_2.editingFinished.connect(self.Manual_Properties)
        self.specific_heat_cv_2.editingFinished.connect(self.Manual_Properties)
        self.compressibility_factor_2.editingFinished.connect(self.Manual_Properties)
        self.molecular_weight_2.editingFinished.connect(self.Manual_Properties)
        self.density_2.editingFinished.connect(self.Manual_Properties)
        self.viscosity_2.editingFinished.connect(self.Manual_Properties)
# End of fluid components and compositions section ************************************************************************

# Start of datasheet and results section **********************************************************************************

        # function called for value generation of vapour conditions
        self.nist_properties_generation.clicked.connect(self.Vapour_Conditions)
        self.fluid_phase_2.activated[str].connect(self.Vapour_Conditions)
        self.specific_heat_cp_2.textChanged.connect(self.Vapour_Conditions)
        self.specific_heat_cv_2.textChanged.connect(self.Vapour_Conditions)
        self.nist_properties_generation.clicked.connect(self.Vapour_Conditions_2)
        self.fluid_phase_2.activated[str].connect(self.Vapour_Conditions_2)
        self.specific_heat_cp_2.editingFinished.connect(self.Vapour_Conditions_2)
        self.specific_heat_cv_2.editingFinished.connect(self.Vapour_Conditions_2)
        self.k_2.editingFinished.connect(self.Vapour_Conditions_2)
        self.pcf_2.editingFinished.connect(self.Vapour_Conditions_2)
        self.vapour_state_2.activated[str].connect(self.Vapour_Conditions_2)

        # function called for calculating coefficient of vapour critical/non-critical flow
        self.tabWidget.currentChanged.connect(self.Vapour_Coefficient_Flow)
        self.vapour_state_2.activated[str].connect(self.Vapour_Coefficient_Flow)

        # function called for calculating Kd (for vapour phase)
        self.rupture_disk_2.activated[str].connect(self.Vapour_Kd_Calc)
        self.vapour_state_2.activated[str].connect(self.Vapour_Kd_Calc)
        self.nist_properties_generation.clicked.connect(self.Vapour_Kd_Calc)
        self.Kd_2.editingFinished.connect(self.Vapour_Kd_Calc_2)

        # function called for calculating Kb (for vapour phase)
        self.design_type_2.activated[str].connect(self.Vapour_Kb_Calc)
        self.vapour_state_2.activated[str].connect(self.Vapour_Kb_Calc)
        self.nist_properties_generation.clicked.connect(self.Vapour_Kb_Calc)
        self.vapour_state_2.activated[str].connect(self.Vapour_Kb_Calc_2)
        self.Kb_2.editingFinished.connect(self.Vapour_Kb_Calc_2)

        # function called for calculating Kc (for vapour phase)
        self.tabWidget.currentChanged.connect(self.Vapour_Kc_Calc)
        self.vapour_state_2.activated[str].connect(self.Vapour_Kc_Calc)
        self.Kc_2.editingFinished.connect(self.Vapour_Kc_Calc_2)

        # function for balanced PRVs (for vapour phase)
        self.tabWidget.currentChanged.connect(self.Balanced_PRV)
        self.design_type_2.activated[str].connect(self.Balanced_PRV)
        self.vapour_state_2.activated[str].connect(self.Balanced_PRV)
        self.mass_flow_2.editingFinished.connect(self.Balanced_PRV)
        self.critical_flow_coefficient_2.editingFinished.connect(self.Balanced_PRV)
        self.subcritical_flow_coefficient_2.editingFinished.connect(self.Balanced_PRV)
        self.Kd_2.editingFinished.connect(self.Balanced_PRV)
        self.relieving_pressure_2.editingFinished.connect(self.Balanced_PRV)
        self.Kb_2.editingFinished.connect(self.Balanced_PRV)
        self.Kc_2.editingFinished.connect(self.Balanced_PRV)
        self.relieving_temp_2.editingFinished.connect(self.Balanced_PRV)
        self.compressibility_factor_2.editingFinished.connect(self.Balanced_PRV)
        self.molecular_weight_2.editingFinished.connect(self.Balanced_PRV)
        self.total_back_pressure_2.editingFinished.connect(self.Balanced_PRV)

        # function for calculating area (for vapour phase)
        self.tabWidget.currentChanged.connect(self.Vapour_Area_Calc)
        self.vapour_state_2.activated[str].connect(self.Vapour_Area_Calc)
        self.mass_flow_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.critical_flow_coefficient_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.subcritical_flow_coefficient_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.Kd_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.relieving_pressure_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.Kb_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.Kc_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.relieving_temp_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.compressibility_factor_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.molecular_weight_2.editingFinished.connect(self.Vapour_Area_Calc)
        self.total_back_pressure_2.editingFinished.connect(self.Vapour_Area_Calc)

        # Function for calculating value generation of steam conditions
        self.tabWidget.currentChanged.connect(self.Steam_Conditions)
        self.fluid_phase_2.activated[str].connect(self.Steam_Conditions)
        self.nist_properties_generation.clicked.connect(self.Steam_Conditions)
        self.KN_2.editingFinished.connect(self.Steam_Conditions_2)
        self.KSH_2.editingFinished.connect(self.Steam_Conditions_2)
        self.relieving_temp_2.editingFinished.connect(self.Steam_Conditions_2)
        self.set_pressure_2.editingFinished.connect(self.Steam_Conditions_2)

        # Function for calculating Kd for steam conditions
        self.rupture_disk_2.activated[str].connect(self.Steam_Kd_Calc)
        self.vapour_state_2.activated[str].connect(self.Steam_Kd_Calc)
        self.nist_properties_generation.clicked.connect(self.Steam_Kd_Calc)
        self.Kd_2.editingFinished.connect(self.Steam_Kd_Calc_2)

        # Functions for calculating Kb for steam conditions
        self.design_type_2.activated[str].connect(self.Steam_Kb_Calc)
        self.vapour_state_2.activated[str].connect(self.Steam_Kb_Calc)
        self.nist_properties_generation.clicked.connect(self.Steam_Kb_Calc)
        self.Kb_2.editingFinished.connect(self.Steam_Kb_Calc_2)

        # Functions called for calculating Kc for steam conditions
        self.rupture_disk_2.activated[str].connect(self.Steam_Kc_Calc)
        self.vapour_state_2.activated[str].connect(self.Steam_Kc_Calc)
        self.nist_properties_generation.clicked.connect(self.Steam_Kc_Calc)
        self.Kc_2.editingFinished.connect(self.Steam_Kc_Calc_2)

        # Functions called for calculating area (for steam conditions)
        self.mass_flow_2.editingFinished.connect(self.Steam_Area_Calc)
        self.relieving_pressure_2.editingFinished.connect(self.Steam_Area_Calc)
        self.Kd_2.editingFinished.connect(self.Steam_Area_Calc)
        self.Kb_2.editingFinished.connect(self.Steam_Area_Calc)
        self.Kc_2.editingFinished.connect(self.Steam_Area_Calc)
        self.KN_2.editingFinished.connect(self.Steam_Area_Calc)
        self.KSH_2.editingFinished.connect(self.Steam_Area_Calc)

        # Function for activating liquid phase calculations (for liquid conditions with capacity certification)
        self.nist_properties_generation.clicked.connect(self.Liquid_Activation)
        self.fluid_phase_2.activated[str].connect(self.Liquid_Activation)

        # Function called for when capacity certification is required (for liquid conditions with capacity certification)
        self.design_type_2.activated[str].connect(self.Capacity_Certification_Yes)
        self.fluid_phase_2.activated[str].connect(self.Capacity_Certification_Yes)
        self.capacity_certification_requirement_2.activated[str].connect(self.Capacity_Certification_Yes)
        self.nist_properties_generation.clicked.connect(self.Capacity_Certification_Yes)
        self.rupture_disk_2.activated[str].connect(self.Capacity_Certification_Yes)
        self.Kw_2.editingFinished.connect(self.Capacity_Certification_Yes_2)
        self.Kd_2.editingFinished.connect(self.Capacity_Certification_Yes_2)
        self.Kc_2.editingFinished.connect(self.Capacity_Certification_Yes_2)

        # Function called for calculating volumetric flow from mass flow (for liquid conditions)
        self.mass_flow_2.editingFinished.connect(self.Vol_Flow_Calc)
        self.density_2.editingFinished.connect(self.Vol_Flow_Calc)

        # Function called for when calculating discharge area (for liquid conditions with capacity certification)
        self.vol_flow_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.Kd_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.Kw_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.Kc_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.Kv_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.density_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.relieving_pressure_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.total_back_pressure_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.viscosity_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.nist_properties_generation.clicked.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.tabWidget.currentChanged.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.mass_flow_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)
        self.vol_flow_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert)

        # Function called for when calculating discharge area (for liquid conditions with capacity certification)
        self.Kw_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert_2)
        self.mass_flow_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert_2)
        self.density_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert_2)
        self.viscosity_2.editingFinished.connect(self.Liquid_Area_Calc_Capacity_Cert_2)
        self.fluid_phase_2.activated[str].connect(self.Liquid_Area_Calc_Capacity_Cert_2)
# End of datasheet and results section ************************************************************************************

# Start of datasheet generation ************************************************************************************
        self.generate_datasheet.clicked.connect(self.Document_Information)
        self.generate_datasheet.clicked.connect(self.General_Equipment_Information)
        self.generate_datasheet.clicked.connect(self.Selection_Basis)
        self.generate_datasheet.clicked.connect(self.Valve_Design)
        self.generate_datasheet.clicked.connect(self.Connections)
        self.generate_datasheet.clicked.connect(self.Materials)
        self.generate_datasheet.clicked.connect(self.Accessories)
        self.generate_datasheet.clicked.connect(self.Service_Conditions)
        self.generate_datasheet.clicked.connect(self.Sizing_And_Selection)
        self.generate_datasheet.clicked.connect(self.Save_Datasheet)
# End of datasheet generation ************************************************************************************


    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "ChemTech - PRV Designer"))
        self.connections.setTitle(_translate("MainWindow", "Connections"))
        self.inlet_size.setText(_translate("MainWindow", "Inlet Size:"))
        self.inlet_size_2.setPlaceholderText(_translate("MainWindow", "Input size here"))
        self.outlet_size.setText(_translate("MainWindow", "Outlet Size:"))
        self.outlet_size_2.setPlaceholderText(_translate("MainWindow", "Input size here"))
        self.inlet_facing.setText(_translate("MainWindow", "Inlet Facing:"))
        self.inlet_facing_2.setItemText(0, _translate("MainWindow", "Raised Face (RF)"))
        self.inlet_facing_2.setItemText(1, _translate("MainWindow", "Flat Face (FF)"))
        self.inlet_facing_2.setItemText(2, _translate("MainWindow", "Ring-Type Joint (RTJ)"))
        self.inlet_facing_2.setItemText(3, _translate("MainWindow", "Tongue-and-Groove (T&G)"))
        self.inlet_facing_2.setItemText(4, _translate("MainWindow", "Male-and-Female (M&F)"))
        self.inlet_facing_2.setItemText(5, _translate("MainWindow", "Other"))
        self.outlet_facing.setText(_translate("MainWindow", "Outlet Facing:"))
        self.outlet_facing_2.setItemText(0, _translate("MainWindow", "Raised Face (RF)"))
        self.outlet_facing_2.setItemText(1, _translate("MainWindow", "Flat Face (FF)"))
        self.outlet_facing_2.setItemText(2, _translate("MainWindow", "Ring-Type Joint (RTJ)"))
        self.outlet_facing_2.setItemText(3, _translate("MainWindow", "Tongue-and-Groove (T&G)"))
        self.outlet_facing_2.setItemText(4, _translate("MainWindow", "Male-and-Female (M&F)"))
        self.outlet_facing_2.setItemText(5, _translate("MainWindow", "Other"))
        self.inlet_facing_3.setPlaceholderText(_translate("MainWindow", "Specify if other"))
        self.outlet_facing_3.setPlaceholderText(_translate("MainWindow", "Specify if other"))
        self.selection_basis.setTitle(_translate("MainWindow", "Selection Basis"))
        self.asme_code.setText(_translate("MainWindow", "ASME Sec VIII Code:"))
        self.asme_code_2.setItemText(0, _translate("MainWindow", "Yes"))
        self.asme_code_2.setItemText(1, _translate("MainWindow", "Yes with U stamp"))
        self.asme_code_2.setItemText(2, _translate("MainWindow", "No"))
        self.asme_code_3.setPlaceholderText(_translate("MainWindow", "Specify if no"))
        self.api_code.setText(_translate("MainWindow", "API 526 Code:"))
        self.api_code_2.setItemText(0, _translate("MainWindow", "Yes"))
        self.api_code_2.setItemText(1, _translate("MainWindow", "No"))
        self.api_code_3.setPlaceholderText(_translate("MainWindow", "Specify if no"))
        self.fire_condition.setText(_translate("MainWindow", "Fire Condition:"))
        self.fire_condition_2.setItemText(0, _translate("MainWindow", "No"))
        self.fire_condition_2.setItemText(1, _translate("MainWindow", "Yes"))
        self.rupture_disk.setText(_translate("MainWindow", "Rupture Disk Present:"))
        self.rupture_disk_2.setItemText(0, _translate("MainWindow", "No"))
        self.rupture_disk_2.setItemText(1, _translate("MainWindow", "Yes"))
        self.equipment_info.setTitle(_translate("MainWindow", "General Equipment Information (optional)"))
        self.item_number.setText(_translate("MainWindow", "Item Number:"))
        self.item_number_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.service_line_number.setText(_translate("MainWindow", "Service Line Number:"))
        self.service_line_number_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.tag_number.setText(_translate("MainWindow", "Tag Number:"))
        self.tag_number_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.number_required.setText(_translate("MainWindow", "Number Required:"))
        self.number_required_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.valve_design.setTitle(_translate("MainWindow", "Valve Design"))
        self.seat_tightness.setText(_translate("MainWindow", "Seat Tightness:"))
        self.nozzle_type.setText(_translate("MainWindow", "Nozzle Type"))
        self.bonnet_type_2.setItemText(0, _translate("MainWindow", "Open"))
        self.bonnet_type_2.setItemText(1, _translate("MainWindow", "Closed"))
        self.nozzle_type_3.setPlaceholderText(_translate("MainWindow", "Specify if other"))
        self.seat_tightness_3.setPlaceholderText(_translate("MainWindow", "Specify if other"))
        self.design_type.setText(_translate("MainWindow", "Design Type:"))
        self.nozzle_type_2.setItemText(0, _translate("MainWindow", "Full"))
        self.nozzle_type_2.setItemText(1, _translate("MainWindow", "Semi"))
        self.nozzle_type_2.setItemText(2, _translate("MainWindow", "Other"))
        self.design_type_2.setItemText(0, _translate("MainWindow", "Conventional"))
        self.design_type_2.setItemText(1, _translate("MainWindow", "Balanced Bellows"))
        self.design_type_2.setItemText(2, _translate("MainWindow", "Balanced Piston"))
        self.bonnet_type.setText(_translate("MainWindow", "Bonnet Type:"))
        self.seat_tightness_2.setItemText(0, _translate("MainWindow", "API 527"))
        self.seat_tightness_2.setItemText(1, _translate("MainWindow", "Other"))
        self.relief_device_role.setText(_translate("MainWindow", "Relief Device Application"))
        self.relief_device_role_2.setItemText(0, _translate("MainWindow", "Single device"))
        self.relief_device_role_2.setItemText(1, _translate("MainWindow", "Additional device"))
        self.relief_device_role_2.setItemText(2, _translate("MainWindow", "Supplemental device"))
        self.materials.setTitle(_translate("MainWindow", "Materials"))
        self.body.setText(_translate("MainWindow", "Body:"))
        self.body_2.setPlaceholderText(_translate("MainWindow", "Specify body material here"))
        self.bonnet.setText(_translate("MainWindow", "Bonnet:"))
        self.bonnet_2.setPlaceholderText(_translate("MainWindow", "Specify bonnet material here"))
        self.seat.setText(_translate("MainWindow", "Seat (Nozzle):"))
        self.seat_2.setPlaceholderText(_translate("MainWindow", "specify seat material here"))
        self.disk.setText(_translate("MainWindow", "Disk"))
        self.disk_2.setPlaceholderText(_translate("MainWindow", "Specify disk material here"))
        self.resilient_seat.setText(_translate("MainWindow", "Resilient Seat:"))
        self.resilient_seat_2.setPlaceholderText(_translate("MainWindow", "Specify resilient seat material here"))
        self.guide.setText(_translate("MainWindow", "Guide:"))
        self.guide_2.setPlaceholderText(_translate("MainWindow", "Specify guide material here"))
        self.adjusting_rings.setText(_translate("MainWindow", "Adjusting Ring(s):"))
        self.adjusting_rings_2.setPlaceholderText(_translate("MainWindow", "Specify adjusting ring(s) material here"))
        self.spring.setText(_translate("MainWindow", "Spring:"))
        self.spring_2.setPlaceholderText(_translate("MainWindow", "Specify spring material here"))
        self.bellows.setText(_translate("MainWindow", "Bellows:"))
        self.bellows_2.setPlaceholderText(_translate("MainWindow", "Specify bellows material here"))
        self.balanced_piston.setText(_translate("MainWindow", "Balanced Pistion:"))
        self.balanced_piston_2.setPlaceholderText(_translate("MainWindow", "Specify balanced piston material"))
        self.nace.setText(_translate("MainWindow", "Comply with NACE:"))
        self.nace_2.setItemText(0, _translate("MainWindow", "No"))
        self.nace_2.setItemText(1, _translate("MainWindow", "Yes"))
        self.internal_gasket.setText(_translate("MainWindow", "Internal Gasket:"))
        self.internal_gasket_2.setPlaceholderText(_translate("MainWindow", "Specify interal gasket material here:"))
        self.document_info.setTitle(_translate("MainWindow", "Document Information (optional)"))
        self.sheet_number.setText(_translate("MainWindow", "Sheet Number:"))
        self.sheet_number_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.job_number.setText(_translate("MainWindow", "Job Number:"))
        self.job_number_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.requisition_number.setText(_translate("MainWindow", "Requisition Number: "))
        self.requisition_number_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.date.setText(_translate("MainWindow", "Date:"))
        self.revision.setText(_translate("MainWindow", "Revision"))
        self.revision_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.by.setText(_translate("MainWindow", "By:"))
        self.by_2.setPlaceholderText(_translate("MainWindow", "Input text here"))
        self.accessories.setTitle(_translate("MainWindow", "Accessories"))
        self.cap.setText(_translate("MainWindow", "Cap:"))
        self.cap_2.setItemText(0, _translate("MainWindow", "Screwed"))
        self.cap_2.setItemText(1, _translate("MainWindow", "Bolted"))
        self.test_gag.setText(_translate("MainWindow", "Test Gag:"))
        self.test_gag_2.setItemText(0, _translate("MainWindow", "Yes"))
        self.test_gag_2.setItemText(1, _translate("MainWindow", "No"))
        self.lifting_lever.setText(_translate("MainWindow", "Lifting Lever:"))
        self.lifting_lever_2.setItemText(0, _translate("MainWindow", "Plain"))
        self.lifting_lever_2.setItemText(1, _translate("MainWindow", "Packed"))
        self.lifting_lever_2.setItemText(2, _translate("MainWindow", "None"))
        self.bug_screen.setText(_translate("MainWindow", "Bug Screen"))
        self.bug_screen_2.setItemText(0, _translate("MainWindow", "Yes"))
        self.bug_screen_2.setItemText(1, _translate("MainWindow", "No"))
        self.bug_screen_2.setItemText(2, _translate("MainWindow", "Other"))
        self.bug_screen_3.setPlaceholderText(_translate("MainWindow", "Specify if other"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "General Information"))
        self.unit_basis.setTitle(_translate("MainWindow", "Unit Basis"))
        self.unit_basis_2.setText(_translate("MainWindow", "Select the unit basis calculations (molar or mass)"))
        self.unit_basis_3.setItemText(0, _translate("MainWindow", "Mass"))
        self.unit_basis_3.setItemText(1, _translate("MainWindow", "Molar"))
        self.fluid_components.setTitle(_translate("MainWindow", "Fluid Components and Compositions"))
        self.fluid_properties_composition.setText(_translate("MainWindow", "Mass Composition"))
        self.fluid_properties_name.setText(_translate("MainWindow", "Fluid Name"))
        self.component_11.setText(_translate("MainWindow", "Component 11"))
        self.component_2_nist.setText(_translate("MainWindow", "-"))
        self.component_1_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_4.setText(_translate("MainWindow", "Component 4"))
        self.component_8.setText(_translate("MainWindow", "Component 8"))
        self.component_12.setText(_translate("MainWindow", "Component 12"))
        self.component_13.setText(_translate("MainWindow", "Component 13"))
        self.component_10.setText(_translate("MainWindow", "Component 10"))
        self.fluid_properties_nist.setText(_translate("MainWindow", "NIST \n"
" REFPROP"))
        self.component_9.setText(_translate("MainWindow", "Component 9"))
        self.component_1_nist.setText(_translate("MainWindow", "-"))
        self.fluid_properties_nist_identification.setText(_translate("MainWindow", "NIST \n"
" Identification"))
        self.component_2_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_7.setText(_translate("MainWindow", "Component 7"))
        self.component_5.setText(_translate("MainWindow", "Component 5"))
        self.component_6.setText(_translate("MainWindow", "Component 6"))
        self.component_1.setText(_translate("MainWindow", "Component 1"))
        self.component_14.setText(_translate("MainWindow", "Component 14"))
        self.component_2.setText(_translate("MainWindow", "Component 2"))
        self.component_3.setText(_translate("MainWindow", "Component 3"))
        self.component_15.setText(_translate("MainWindow", "Component 15"))
        self.component_3_nist.setText(_translate("MainWindow", "-"))
        self.component_4_nist.setText(_translate("MainWindow", "-"))
        self.component_5_nist.setText(_translate("MainWindow", "-"))
        self.component_6_nist.setText(_translate("MainWindow", "-"))
        self.component_7_nist.setText(_translate("MainWindow", "-"))
        self.component_8_nist.setText(_translate("MainWindow", "-"))
        self.component_9_nist.setText(_translate("MainWindow", "-"))
        self.component_10_nist.setText(_translate("MainWindow", "-"))
        self.component_11_nist.setText(_translate("MainWindow", "-"))
        self.component_12_nist.setText(_translate("MainWindow", "-"))
        self.component_13_nist.setText(_translate("MainWindow", "-"))
        self.component_14_nist.setText(_translate("MainWindow", "-"))
        self.component_15_nist.setText(_translate("MainWindow", "-"))
        self.component_3_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_4_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_5_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_6_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_7_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_8_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_9_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_10_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_11_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_12_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_13_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_14_nist_identification.setText(_translate("MainWindow", "-"))
        self.component_15_nist_identification.setText(_translate("MainWindow", "-"))
        self.total_composition.setText(_translate("MainWindow", "Total Composition"))
        self.conditions.setTitle(_translate("MainWindow", "Conditions"))
        self.total_back_pressure.setText(_translate("MainWindow", "Total Back Pressure P2 (barg)"))
        self.relieving_pressure_2.setPlaceholderText(_translate("MainWindow", "Enter relieving pressure here"))
        self.set_pressure_2.setPlaceholderText(_translate("MainWindow", "Enter pressure here"))
        self.relieving_temp_2.setPlaceholderText(_translate("MainWindow", "Enter temperature here"))
        self.overpressure_percentage_2.setPlaceholderText(_translate("MainWindow", "Enter over pressure here"))
        self.total_back_pressure_2.setPlaceholderText(_translate("MainWindow", "Enter back pressure here"))
        self.set_pressure.setText(_translate("MainWindow", "PRV Set Pressure (barg)"))
        self.relieving_temp.setText(_translate("MainWindow", "Relieving Temperature (°C)"))
        self.relieving_pressure.setText(_translate("MainWindow", "Relieving Pressure P1 (barg)"))
        self.overpressure_percentage.setText(_translate("MainWindow", "Maximum Allowable Accumulated Pressure (%)"))
        self.equipment_mawp.setText(_translate("MainWindow", "Equipment MAWP (barg)"))
        self.equipment_mawp_2.setPlaceholderText(_translate("MainWindow", "Enter pressure here"))
        self.fluid_properties_results.setTitle(_translate("MainWindow", "Fluid Properties"))
        self.molecular_weight_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.density_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.specific_heat_cp_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.density.setText(_translate("MainWindow", "Density ρ (kg/m3)"))
        self.compressibility_factor.setText(_translate("MainWindow", "Compressibility Factor Z"))
        self.specific_heat_cp.setText(_translate("MainWindow", "Specific heat Cp (J/(kg K))"))
        self.compressibility_factor_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.fluid_phase_2.setCurrentText(_translate("MainWindow", "Select manually here"))
        self.fluid_phase_2.setItemText(0, _translate("MainWindow", "Select manually here"))
        self.fluid_phase_2.setItemText(1, _translate("MainWindow", "Vapour"))
        self.fluid_phase_2.setItemText(2, _translate("MainWindow", "Liquid"))
        self.fluid_phase_2.setItemText(3, _translate("MainWindow", "Liquid/Vapour"))
        self.fluid_phase_2.setItemText(4, _translate("MainWindow", "Supercritical"))
        self.fluid_phase.setText(_translate("MainWindow", "Fluid Phase"))
        self.molecular_weight.setText(_translate("MainWindow", "Molecular Weight M (kg/kg-mol)"))
        self.specific_heat_cv.setText(_translate("MainWindow", "Specific heat Cv (J/(kg K))"))
        self.specific_heat_cv_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.viscosity.setText(_translate("MainWindow", "Viscosity μ (cP)"))
        self.viscosity_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.nist_properties_generation.setText(_translate("MainWindow", "Generate Properties From NIST Refprop"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Fluid Properties"))
        self.generate_datasheet.setText(_translate("MainWindow", "Generate Datasheet"))
        self.variables.setTitle(_translate("MainWindow", "Variables"))
        self.Kc.setText(_translate("MainWindow", "Combination Correction Factor Kc"))
        self.Kb.setText(_translate("MainWindow", "Capacity Correction Factor Kb"))
        self.mass_flow.setText(_translate("MainWindow", "Required Mass Flow W (kg/h)"))
        self.mass_flow_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.Kb_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.Kd_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.Kc_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.Kd.setText(_translate("MainWindow", "Effective Coefficient of discharge Kd"))
        self.vol_flow.setText(_translate("MainWindow", "Required Volumetric Flow (L/min)"))
        self.vol_flow_2.setPlaceholderText(_translate("MainWindow", "Input manually here"))
        self.steam_conditions.setTitle(_translate("MainWindow", "Steam Conditions"))
        self.KN.setText(_translate("MainWindow", "Correction Factor for Napier Equation KN"))
        self.KSH.setText(_translate("MainWindow", "Superheat Correction Factor KSH"))
        self.liquid_phase_conditions.setTitle(_translate("MainWindow", "Liquid Phase Conditions"))
        self.Kp.setText(_translate("MainWindow", "Correction Factor due to Overpressure Kp"))
        self.capacity_certification_requirement_2.setItemText(0, _translate("MainWindow", "Yes"))
        self.capacity_certification_requirement_2.setItemText(1, _translate("MainWindow", "No"))
        self.capacity_certification_requirement.setText(_translate("MainWindow", "Capacity Certification Requirement"))
        self.Kw.setText(_translate("MainWindow", "Correction Factor due to Backpressure Kw"))
        self.Kv.setText(_translate("MainWindow", "Correction Factor due to Viscosity Kv"))
        self.Re.setText(_translate("MainWindow", "Reynolds Number Re"))
        self.vapour_conditions.setTitle(_translate("MainWindow", "Vapour Conditions"))
        self.pcf.setText(_translate("MainWindow", "Critical Flow Pressure Pcf (bar)"))
        self.critical_flow_coefficient.setText(_translate("MainWindow", "Coefficient of Critical Flow C"))
        self.vapour_state.setText(_translate("MainWindow", "Vapour Flow State"))
        self.k.setText(_translate("MainWindow", "Ratio of Specific Heats k"))
        self.vapour_state_2.setItemText(0, _translate("MainWindow", "Edit manually here"))
        self.vapour_state_2.setItemText(1, _translate("MainWindow", "Critical"))
        self.vapour_state_2.setItemText(2, _translate("MainWindow", "Non-critical"))
        self.subcritical_flow_coefficient.setText(_translate("MainWindow", "Coefficient of Subcritical Flow F2"))
        self.warning_error_messages.setTitle(_translate("MainWindow", "Warnings and Error Messages"))
        self.results.setTitle(_translate("MainWindow", "Results"))
        self.discharge_area.setText(_translate("MainWindow", "Effective Discharge Area (mm2)"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("MainWindow", "Datasheet Generation"))

# Start of UI Code ----------------------------------------------------------------------------------------------------------------------------------
    #asme code combobox UI logic
    def Asme_Code_UI_Logic(self):
        if self.asme_code_2.currentText() == 'No':
            self.asme_code_3.setEnabled(True)
            self.asme_code_3.setReadOnly(False)
        elif self.asme_code_2.currentText() == 'Yes':
            self.asme_code_3.setEnabled(False)
            self.asme_code_3.setReadOnly(True)
        elif self.asme_code_2.currentText() == 'Yes with U stamp':
            self.asme_code_3.setEnabled(False)
            self.asme_code_3.setReadOnly(True)

    #api 526 combobox UI logic
    def API_Code_UI_Logic(self):
        if self.api_code_2.currentText() == 'No':
            self.api_code_3.setEnabled(True)
            self.api_code_3.setReadOnly(False)
        elif self.api_code_2.currentText() == 'Yes':
            self.api_code_3.setEnabled(False)
            self.api_code_3.setReadOnly(True)

    #nozzle stype UI logic
    def Nozzle_Type_Logic(self):
        if self.nozzle_type_2.currentText() == 'Other':
            self.nozzle_type_3.setEnabled(True)
            self.nozzle_type_3.setReadOnly(False)
        elif self.nozzle_type_2.currentText() == 'Semi':
            self.nozzle_type_3.setEnabled(False)
            self.nozzle_type_3.setReadOnly(True)
        elif self.nozzle_type_2.currentText() == 'Full':
            self.nozzle_type_3.setEnabled(False)
            self.nozzle_type_3.setReadOnly(True)

    #seat tightness UI logic
    def Seat_Tightness(self):
        if self.seat_tightness_2.currentText() == 'Other':
            self.seat_tightness_3.setEnabled(True)
            self.seat_tightness_3.setReadOnly(False)
        elif self.seat_tightness_2.currentText() == 'API 527':
            self.seat_tightness_3.setEnabled(False)
            self.seat_tightness_3.setReadOnly(True)

    #inlet flange rating UI logic
    def Inlet_Facing(self):
        if self.inlet_facing_2.currentText() == 'Other':
            self.inlet_facing_3.setEnabled(True)
            self.inlet_facing_3.setReadOnly(False)
        else:
            self.inlet_facing_3.setEnabled(False)
            self.inlet_facing_3.setReadOnly(True)

    #outlet flange rating UI logic
    def Outlet_Facing(self):
        if self.outlet_facing_2.currentText() == 'Other':
            self.outlet_facing_3.setEnabled(True)
            self.outlet_facing_3.setReadOnly(False)
        else:
            self.outlet_facing_3.setEnabled(False)
            self.outlet_facing_3.setReadOnly(True)

    #bug screen UI logic
    def Bug_Screen(self):
        pass
# End of UI Code -------------------------------------------------------------------------------------------------------------------------------------



# Start of fluid components and compositions section -------------------------------------------------------------------------------------------------
    # Base units combobox logic
    def Unit_Basis(self):
        if self.unit_basis_3.currentText() == 'Mass':
            BASE_UNIT_SI = RP.GETENUMdll(0, 'MASS BASE SI').iEnum
            self.specific_heat_cp.setText('Specific heat Cp (J/(kg K))')
            self.specific_heat_cv.setText('Specific heat Cv (J/(kg K))')
            self.density.setText('Density ρ (kg/m3)')
            self.fluid_properties_composition.setText('Mass Composition')
        elif self.unit_basis_3.currentText() == 'Molar':
            BASE_UNIT_SI = RP.GETENUMdll(0, 'MOLAR BASE SI').iEnum
            self.specific_heat_cp.setText('Specific heat Cp (J/(mol K))')
            self.specific_heat_cv.setText('Specific heat Cv (J/(mol K))')
            self.density.setText('Density ρ (mol/m3)')
            self.fluid_properties_composition.setText('Molar Composition')

    # Component name comparison with NIST logic
    def Component_NIST_Check(self):
        for i in range(1, 16):
            name = getattr(self, 'component_{}_name'.format(i)).text()
            NIST_name = RP.REFPROPdll(name, '', 'NAME', BASE_UNIT_SI, 0, 0, 0, 0, [1])
            if name == '':
                getattr(self, 'component_{}_nist'.format(i)).setText('-')
                getattr(self, 'component_{}_nist_identification'.format(i)).setText('-')
            elif NIST_name.hUnits == '':
                getattr(self, 'component_{}_nist'.format(i)).setText('No')
                getattr(self, 'component_{}_nist_identification'.format(i)).setText('-')
            else:
                getattr(self, 'component_{}_nist'.format(i)).setText('Yes')
                getattr(self, 'component_{}_nist_identification'.format(i)).setText(NIST_name.hUnits)

    #component name repeat check
    def Component_Name_Repeat_Check(self):
        for i in range(1, 16):
            component = getattr(self, 'component_{}_name'.format(i)).text().casefold()
            for x in range(1, 16):
                if getattr(self, 'component_{}_name'.format(x)).text() == '':
                    pass
                elif str(getattr(self, 'component_{}_name'.format(x))).casefold() == str(getattr(self, 'component_{}_name'.format(i))).casefold():
                    pass
                elif str(getattr(self, 'component_{}_name'.format(x)).text()).casefold() == component:
                    repeat_name_error = QMessageBox()
                    repeat_name_error.setWindowTitle('ChemTech - Error')
                    repeat_name_error.setText('Cannot repeat component names')
                    repeat_name_error.setIcon(QMessageBox.Critical)
                    error_message = repeat_name_error.exec_()
                    getattr(self, 'component_{}_name'.format(x)).setText('')
                    break
        for i in range(2,16): #doing check to see if the next component number is filled in
            component = getattr(self, 'component_{}_name'.format(i)).text()
            if component != '' and getattr(self, 'component_{}_name'.format(i-1)).text() == '':
                fill_next_comp_number = QMessageBox()
                fill_next_comp_number.setWindowTitle('ChemTech - Error')
                fill_next_comp_number.setText('Please fill in the next component number instead')
                fill_next_comp_number.setIcon(QMessageBox.Critical)
                error_message = fill_next_comp_number.exec_()
                getattr(self, 'component_{}_name'.format(i)).setText('')
                break
            else:
                pass

    # component composition logic
    def Composition_Number_Test(self):
        for i in range(2,16): #doing check to see if the next component number is filled in
            component = getattr(self, 'component_{}_composition'.format(i)).text()
            if component != '' and getattr(self, 'component_{}_composition'.format(i-1)).text() == '':
                fill_next_comp_number = QMessageBox()
                fill_next_comp_number.setWindowTitle('ChemTech - Error')
                fill_next_comp_number.setText('Please fill in the next component number instead')
                fill_next_comp_number.setIcon(QMessageBox.Critical)
                error_message = fill_next_comp_number.exec_()
                getattr(self, 'component_{}_composition'.format(i)).setText('')
                break
            else:
                pass
        for i in range(1, 16):
            composition = getattr(self, 'component_{}_composition'.format(i)).text()
            if composition.replace('.','',1).isdigit() == True or composition == '':
                pass
            else:
                no_number_error = QMessageBox()
                no_number_error.setWindowTitle('ChemTech - Error')
                no_number_error.setText('Number required between 0 and 1')
                no_number_error.setIcon(QMessageBox.Critical)
                error_message = no_number_error.exec_()
                getattr(self, 'component_{}_composition'.format(i)).setText('')
                break

    # composition sum check
    def Composition_Sum_Test(self):
        for j in range(2, 16):
            if getattr(self, 'component_{}_composition'.format(j)).text() == '' and getattr(self, 'component_{}_composition'.format(j-1)).text() == '':
                self.total_composition_2.setText('')
            else:
                pass
        composition_sum = 0
        for i in range(1, 16):
            if getattr(self, 'component_{}_composition'.format(i)).text() == '':
                pass
            else:
                composition_sum = composition_sum + float(getattr(self, 'component_{}_composition'.format(i)).text())
                composition_sum = round(composition_sum, 4)
                self.total_composition_2.setText(str(composition_sum))
                if composition_sum > 1:
                    comp_sum_error = QMessageBox()
                    comp_sum_error.setWindowTitle('ChemTech - Error')
                    comp_sum_error.setText('Sum of compositions cannot be greater than 1')
                    comp_sum_error.setIcon(QMessageBox.Critical)
                    error_message = comp_sum_error.exec_()
                    #getattr(self, 'component_{}_composition'.format(i)).setText('')
                    self.total_composition_2.setText('ERROR: Compositions > 1 (' + (str(composition_sum)) + ')')
                    break
                else:
                    pass

    # activation of 'generate NIST properties' button
    def NIST_Properties_Generation(self):
        def Is_Number(number):
            try:
                float(number)
            except ValueError:
                return False
            return True
        list = []
        for i in range (1, 16):
            list.append(getattr(self, 'component_{}_nist'.format(i)).text())
        if 'No' in list:
            self.nist_properties_generation.setEnabled(False)
            self.fluid_phase_2.setCurrentText('Select manually here')
            self.fluid_phase_2.setStyleSheet('font: non-italic; color: black')
            self.specific_heat_cp_2.setText('')
            self.specific_heat_cp_2.setStyleSheet('font: non-italic; color: black')
            self.specific_heat_cv_2.setText('')
            self.specific_heat_cv_2.setStyleSheet('font: non-italic; color: black')
            self.compressibility_factor_2.setText('')
            self.compressibility_factor_2.setStyleSheet('font: non-italic; color: black')
            self.molecular_weight_2.setText('')
            self.molecular_weight_2.setStyleSheet('font: non-italic; color: black')
            self.density_2.setText('')
            self.density_2.setStyleSheet('font: non-italic; color: black')
            self.viscosity_2.setText('')
            self.viscosity_2.setStyleSheet('font: non-italic; color: black')
            return None
        else:
            pass
        for i in range (1, 16):
            if getattr(self, 'component_{}_name'.format(i)).text() != '' and getattr(self, 'component_{}_composition'.format(i)).text() != '':
                pass
            elif getattr(self, 'component_{}_name'.format(i)).text() == '' and getattr(self, 'component_{}_composition'.format(i)).text() == '':
                pass
            else:
                self.nist_properties_generation.setEnabled(False)
                return None

        if self.total_composition_2.text() == '':
            self.nist_properties_generation.setEnabled(False)
            return None
        elif Is_Number(self.total_composition_2.text()) == True:
            pass
        elif Is_Number(self.total_composition_2.text()) == False:
            return None
        elif round(float(self.total_composition_2.text()), 4) == 1.0:
            pass
        else:
            self.nist_properties_generation.setEnabled(False)
            return None

        if self.relieving_temp_2.text() != '' and self.relieving_pressure_2.text() != '':
            self.nist_properties_generation.setEnabled(True)
        else:
            self.nist_properties_generation.setEnabled(False)
            return None

    # Conditions section number check
    def Conditions_Number_Check(self):
        dict = {}
        def Is_Number(number):
            try:
                float(number)
            except ValueError:
                return False
        dict[self.relieving_temp_2] = self.relieving_temp_2.text()
        dict[self.equipment_mawp_2] = self.equipment_mawp_2.text()
        dict[self.set_pressure_2] = self.set_pressure_2.text()
        dict[self.overpressure_percentage_2] = self.overpressure_percentage_2.text()
        dict[self.total_back_pressure_2] = self.total_back_pressure_2.text()
        for key, value in dict.items():
            if key == self.set_pressure_2 and self.set_pressure_2.text() == '<None>':
                pass
            elif key == self.overpressure_percentage_2 and self.overpressure_percentage_2.text() == '<None>':
                pass
            elif key != self.set_pressure_2 and value == '':
                pass
            elif key != self.overpressure_percentage_2 and value == '':
                pass
            elif Is_Number(value) == True:
                pass
            elif Is_Number(value) == False:
                no_number_error = QMessageBox()
                no_number_error.setWindowTitle('ChemTech - Error')
                no_number_error.setText('You must only enter a number here to continue')
                no_number_error.setIcon(QMessageBox.Critical)
                error_message = no_number_error.exec_()
                if key == self.set_pressure_2 or key == self.overpressure_percentage_2:
                    key.setText('<None>')
                else:
                    key.setText('')
                break

    # Logic for the set pressure line edit for intial value
    def Equipment_MAWP(self):
        set_pressure = 0
        if self.equipment_mawp_2.text() == '':
            pass
        elif self.relief_device_role_2.currentText() == 'Single device':
            set_pressure = 1 * float(self.equipment_mawp_2.text())
        elif self.relief_device_role_2.currentText() == 'Additional device':
            set_pressure = 1.05 * float(self.equipment_mawp_2.text())
        elif self.relief_device_role_2.currentText() == 'Supplemental device':
            set_pressure = 1.1 * float(self.equipment_mawp_2.text())

        set_pressure = round(set_pressure, 2)
        set_pressure = str(set_pressure)
        if self.equipment_mawp_2.text() != '':
            self.set_pressure_2.setText(set_pressure)
            self.set_pressure_2.setStyleSheet('font: italic; color: blue')
        elif self.equipment_mawp_2.text() == '':
            self.set_pressure_2.setText('')
            self.set_pressure_2.setStyleSheet('font: non-italic; color: black')
        else:
            pass

    # Logic for the set pressure line edit if manually inputted
    def Set_Pressure(self):
        set_pressure = None
        if self.equipment_mawp_2.text() == '':
            pass
        elif self.relief_device_role_2.currentText() == 'Single device':
            set_pressure = 1 * float(self.equipment_mawp_2.text())
        elif self.relief_device_role_2.currentText() == 'Additional device':
            set_pressure = 1.05 * float(self.equipment_mawp_2.text())
        elif self.relief_device_role_2.currentText() == 'Supplemental device':
            set_pressure = 1.1 * float(self.equipment_mawp_2.text())

        if set_pressure != None:
            set_pressure = round(set_pressure, 2)
            set_pressure = str(set_pressure)
        else:
            pass

        if self.set_pressure_2.text() == '':
            pass
        elif self.equipment_mawp_2.text() == '':
            self.set_pressure_2.setStyleSheet('font: italic; color: blue')
        elif float(self.set_pressure_2.text()) == 1 * float(set_pressure):
            self.set_pressure_2.setStyleSheet('font: italic; color: blue')
        elif float(self.set_pressure_2.text()) == 1.05 * float(set_pressure):
            self.set_pressure_2.setStyleSheet('font: italic; color: blue')
        elif float(self.set_pressure_2.text()) == 1.1 * float(set_pressure):
            self.set_pressure_2.setStyleSheet('font: italic; color: blue')
        else:
            self.set_pressure_2.setStyleSheet('font: non-italic; color: black')

        if self.set_pressure_2.text() == '':
            self.set_pressure_2.setText(set_pressure)
            self.set_pressure_2.setStyleSheet('font: italic; color: blue')
        elif self.set_pressure_2.text() == None:
            self.set_pressure_2.setText('')

    # Logic for the overpressure line edit
    def Overpressure_Percentage(self):
        if (self.overpressure_percentage_2.text() == '' or self.overpressure_percentage_2.text() == '110'
        or self.overpressure_percentage_2.text() == '121' or self.overpressure_percentage_2.text() == '116'):
            if self.fire_condition_2.currentText() == 'No' and self.relief_device_role_2.currentText() == 'Single device':
                self.overpressure_percentage_2.setText('110')
                self.overpressure_percentage_2.setStyleSheet('font: italic; color: blue')
            elif self.fire_condition_2.currentText() == 'Yes' and self.relief_device_role_2.currentText() == 'Single device':
                self.overpressure_percentage_2.setText('121')
                self.overpressure_percentage_2.setStyleSheet('font: italic; color: blue')
            elif self.fire_condition_2.currentText() == 'No' and self.relief_device_role_2.currentText() == 'Additional device':
                self.overpressure_percentage_2.setText('116')
                self.overpressure_percentage_2.setStyleSheet('font: italic; color: blue')
            elif self.fire_condition_2.currentText() == 'Yes' and self.relief_device_role_2.currentText() == 'Additional device':
                self.overpressure_percentage_2.setText('121')
                self.overpressure_percentage_2.setStyleSheet('font: italic; color: blue')
            elif self.fire_condition_2.currentText() == 'Yes' and self.relief_device_role_2.currentText() == 'Supplemental device':
                self.overpressure_percentage_2.setText('121')
                self.overpressure_percentage_2.setStyleSheet('font: italic; color: blue')
        else:
            self.overpressure_percentage_2.setStyleSheet('font: non-italic; color: black')

    # Calculation of initial relieving pressure value
    def Relieving_Pressure_Calc(self):
        if self.total_back_pressure_2.text() == '':
            self.relieving_pressure_2.setText('')
            return None
        elif self.equipment_mawp_2.text() == '':
            self.relieving_pressure_2.setText('')
            return None
        elif self.overpressure_percentage_2.text() == '':
            self.relieving_pressure_2.setText('')
            return None
        back_pressure = float(self.total_back_pressure_2.text()) + 1.01325
        mawp = float(self.equipment_mawp_2.text())
        max_accumulated_pressure = (float(self.overpressure_percentage_2.text()) / 100) * mawp
        relieving_pressure = (max_accumulated_pressure + back_pressure) - 1.01325
        relieving_pressure = str(round(relieving_pressure, 2))
        self.relieving_pressure_2.setText(relieving_pressure)
        self.relieving_pressure_2.setStyleSheet('font: italic; color: blue')
        # reseting fluid property values if above conditions are changed
        self.fluid_phase_2.setCurrentText('Select manually here')
        self.specific_heat_cp_2.setText('')
        self.specific_heat_cv_2.setText('')
        self.compressibility_factor_2.setText('')
        self.molecular_weight_2.setText('')
        self.density_2.setText('')
        self.viscosity_2.setText('')
    # Caclulation of relieving pressure value
    def Relieving_Pressure_Calc_2(self):
        if self.total_back_pressure_2.text() == '':
            self.relieving_pressure_2.setText('')
            return None
        elif self.equipment_mawp_2.text() == '':
            self.relieving_pressure_2.setText('')
            return None
        elif self.overpressure_percentage_2.text() == '':
            self.relieving_pressure_2.setText('')
            return None
        mawp = float(self.equipment_mawp_2.text())
        overpressure = float(self.overpressure_percentage_2.text()) / 100
        relieving_pressure = overpressure*mawp
        relieving_pressure = str(round(relieving_pressure, 2))
        if self.relieving_pressure_2.text() == '':
            self.relieving_pressure_2.setText(relieving_pressure)
            self.relieving_pressure_2.setStyleSheet('font: italic; color: blue')
        elif self.relieving_pressure_2.text() == relieving_pressure:
            self.relieving_pressure_2.setText(relieving_pressure)
            self.relieving_pressure_2.setStyleSheet('font: italic; color: blue')
        elif self.relieving_pressure_2.text() != relieving_pressure:
            self.relieving_pressure_2.setStyleSheet('font: non-italic; color: black')
        self.fluid_phase_2.setCurrentText('Select manually here')
        self.specific_heat_cp_2.setText('')
        self.specific_heat_cv_2.setText('')
        self.compressibility_factor_2.setText('')
        self.molecular_weight_2.setText('')
        self.density_2.setText('')
        self.viscosity_2.setText('')

    # Initial generation of NIST properties
    def NIST_Properties(self):
        comps = []
        names = ''
        for i in range(1, 16):
            if (getattr(self, 'component_{}_name'.format(i)).text() != ''
            and getattr(self, 'component_{}_composition'.format(i)).text() != ''):
                names = names + getattr(self, 'component_{}_name'.format(i)).text() + ' * '
                comps.append(float(getattr(self, 'component_{}_composition'.format(i)).text()))
            else:
                pass
        names = names[:-3]
        temp = float(self.relieving_temp_2.text()) + 273.15
        pressure = (float(self.relieving_pressure_2.text()) + 1.01325) * 100000
        Units = []
        Results = []
        for k in ['PHASE', 'Cp', 'Cv', 'Z', 'M', 'D', 'VIS']:
            if self.unit_basis_3.currentText() == 'Mass':
                props = RP.REFPROPdll(names, 'TP', k, MASS_UNIT_SI, 0, 0, temp, pressure, comps)
                Units.append(props.hUnits)
                Results.append(props.Output[0])
            elif self.unit_basis_3.currentText() == 'Molar':
                props = RP.REFPROPdll(names, 'TP', k, MOLAR_UNIT_SI, 0, 0, temp, pressure, comps)
                Units.append(props.hUnits)
                Results.append(props.Output[0])
        phase = Units[0]
        if phase == 'Subcooled liquid':
            self.fluid_phase_2.setCurrentText('Liquid')
            self.fluid_phase_2.setStyleSheet('font: italic; color: blue')
        elif phase == 'Superheated gas':
            self.fluid_phase_2.setCurrentText('Vapour')
            self.fluid_phase_2.setStyleSheet('font: italic; color: blue')
        elif phase == 'Two-phase':
            self.fluid_phase_2.setCurrentText('Liquid/Vapour')
            self.fluid_phase_2.setStyleSheet('font: italic; color: blue')
        elif phase == 'Supercritical':
            self.fluid_phase_2.setCurrentText('Supercritical')
            self.fluid_phase_2.setStyleSheet('font: italic; color: blue')
        Results[1] = round(Results[1], 2)
        self.specific_heat_cp_2.setText(str(Results[1]))
        self.specific_heat_cp_2.setStyleSheet('font: italic; color: blue')
        Results[2] = round(Results[2], 2)
        self.specific_heat_cv_2.setText(str(Results[2]))
        self.specific_heat_cv_2.setStyleSheet('font: italic; color: blue')
        Results[3] = round(Results[3], 4)
        self.compressibility_factor_2.setText(str(Results[3]))
        self.compressibility_factor_2.setStyleSheet('font: italic; color: blue')
        Results[4] = Results[4] * 1000
        Results[4] = round(Results[4], 2)
        self.molecular_weight_2.setText(str(Results[4]))
        self.molecular_weight_2.setStyleSheet('font: italic; color: blue')
        Results[5] = round(Results[5], 2)
        self.density_2.setText(str(Results[5]))
        self.density_2.setStyleSheet('font: italic; color: blue')
        Results[6] = '{:f}'.format(Results[6])
        self.viscosity_2.setText(Results[6])
        self.viscosity_2.setStyleSheet('font: italic; color: blue')
        dict = {}
        dict[self.specific_heat_cp_2] = self.specific_heat_cp_2.text()
        dict[self.specific_heat_cv_2] = self.specific_heat_cv_2.text()
        dict[self.compressibility_factor_2] = self.compressibility_factor_2.text()
        dict[self.density_2] = self.density_2.text()
        dict[self.viscosity_2] = self.viscosity_2.text()
        for key, value in dict.items():
            if str(-9999990) in value:
                key.setText('No value found from NIST')
                key.setStyleSheet('font: italic; color: blue')
            else:
                pass

    def Manual_Properties(self):
        if self.relieving_temp_2.text() == '':
            return None
        if self.equipment_mawp_2.text() == '':
            return None
        if self.total_back_pressure_2.text() == '':
            return None
        if self.relieving_pressure_2.text() == '':
            return None
        comps = []
        names = ''
        for i in range(1, 16):
            if (getattr(self, 'component_{}_name'.format(i)).text() != ''
            and getattr(self, 'component_{}_composition'.format(i)).text() != ''):
                names = names + getattr(self, 'component_{}_name'.format(i)).text() + ' * '
                comps.append(float(getattr(self, 'component_{}_composition'.format(i)).text()))
            else:
                pass
        names = names[:-3]
        temp = float(self.relieving_temp_2.text()) + 273.15
        pressure = (float(self.relieving_pressure_2.text()) + 1.01325) * 100000
        Units = []
        Results = []
        for k in ['PHASE', 'Cp', 'Cv', 'Z', 'M', 'D', 'VIS']:
            if self.unit_basis_3.currentText() == 'Mass':
                props = RP.REFPROPdll(names, 'TP', k, MASS_UNIT_SI, 0, 0, temp, pressure, comps)
                Units.append(props.hUnits)
                Results.append(props.Output[0])
            elif self.unit_basis_3.currentText() == 'Molar':
                props = RP.REFPROPdll(names, 'TP', k, MOLAR_UNIT_SI, 0, 0, temp, pressure, comps)
                Units.append(props.hUnits)
                Results.append(props.Output[0])

        if Units[0] == 'Subcooled liquid' and self.fluid_phase_2.currentText() != 'Liquid':
            self.fluid_phase_2.setStyleSheet('font: non-italic; color: black')
        elif Units[0] == 'Superheated gas' and self.fluid_phase_2.currentText() != 'Vapour':
            self.fluid_phase_2.setStyleSheet('font: non-italic; color: black')
        elif Units[0] == 'Two-phase' and self.fluid_phase_2.currentText() != 'Liquid/Vapour':
            self.fluid_phase_2.setStyleSheet('font: non-italic; color: black')
        elif Units[0] == 'Supercritical' and self.fluid_phase_2.currentText() != 'Supercritical':
            self.fluid_phase_2.setStyleSheet('font: non-italic; color: black')
        else:
            self.fluid_phase_2.setStyleSheet('font: italic; color: blue')

        Results[1] = str(round(Results[1], 2))
        if self.specific_heat_cp_2.text() == Results[1] or self.specific_heat_cp_2.text() == 'No value found from NIST':
            self.specific_heat_cp_2.setText(Results[1])
            self.specific_heat_cp_2.setStyleSheet('font: italic; color: blue')
        elif self.specific_heat_cp_2.text().replace('.','',1).isdigit() != True:
            self.specific_heat_cp_2.setText('')
        elif self.specific_heat_cp_2.text() != Results[1]:
            self.specific_heat_cp_2.setStyleSheet('font: non-italic; color: black')

        Results[2] = str(round(Results[2], 2))
        if self.specific_heat_cv_2.text() == Results[2] or self.specific_heat_cv_2.text() == 'No value found from NIST':
            self.specific_heat_cv_2.setText(Results[2])
            self.specific_heat_cv_2.setStyleSheet('font: italic; color: blue')
        elif self.specific_heat_cv_2.text().replace('.','',1).isdigit() != True:
            self.specific_heat_cv_2.setText('')
        elif self.specific_heat_cv_2.text() != Results[2]:
            self.specific_heat_cv_2.setStyleSheet('font: non-italic; color: black')

        Results[3] = str(round(Results[3], 4))
        if self.compressibility_factor_2.text() == Results[3] or self.compressibility_factor_2.text() == 'No value found from NIST':
            self.compressibility_factor_2.setText(Results[3])
            self.compressibility_factor_2.setStyleSheet('font: italic; color: blue')
        elif self.compressibility_factor_2.text().replace('.','',1).isdigit() != True:
            self.compressibility_factor_2.setText('')
        elif self.compressibility_factor_2.text() != Results[3]:
            self.compressibility_factor_2.setStyleSheet('font: non-italic; color: black')

        Results[4] = Results[4] * 1000
        Results[4] = str(round(Results[4], 2))
        if self.molecular_weight_2.text() == Results[4] or self.molecular_weight_2.text() == 'No value found from NIST':
            self.molecular_weight_2.setText(Results[4])
            self.molecular_weight_2.setStyleSheet('font: italic; color: blue')
        elif self.molecular_weight_2.text().replace('.','',1).isdigit() != True:
            self.molecular_weight_2.setText('')
        elif self.molecular_weight_2.text() != Results[4]:
            self.molecular_weight_2.setStyleSheet('font: non-italic; color: black')

        Results[5] = str(round(Results[5], 2))
        if self.density_2.text() == Results[5] or self.density_2.text() == 'No value found from NIST':
            self.density_2.setText(Results[5])
            self.density_2.setStyleSheet('font: italic; color: blue')
        elif self.density_2.text().replace('.','',1).isdigit() != True:
            self.density_2.setText('')
        elif self.density_2.text() != Results[5]:
            self.density_2.setStyleSheet('font: non-italic; color: black')

        Results[6] = '{:f}'.format(Results[6])
        if self.viscosity_2.text() == Results[6] or self.viscosity_2.text() == 'No value found from NIST':
            self.viscosity_2.setText(Results[6])
            self.viscosity_2.setStyleSheet('font: italic; color: blue')
        elif self.viscosity_2.text().replace('.','',1).isdigit() != True:
            self.viscosity_2.setText('')
        elif self.viscosity_2.text() != Results[6]:
            self.viscosity_2.setStyleSheet('font: non-italic; color: black')

        dict = {}
        dict[self.specific_heat_cp_2] = self.specific_heat_cp_2.text()
        dict[self.specific_heat_cv_2] = self.specific_heat_cv_2.text()
        dict[self.compressibility_factor_2] = self.compressibility_factor_2.text()
        dict[self.density_2] = self.density_2.text()
        dict[self.viscosity_2] = self.viscosity_2.text()
        for key, value in dict.items():
            if str(-9999990) in value:
                key.setText('No value found from NIST')
                key.setStyleSheet('font: italic; color: blue')
            else:
                pass
# End of fluid components and composition section ----------------------------------------------------------------------------------------------------


# Start of datasheet and results section -------------------------------------------------------------------------------------------------------------

    #VAPOUR CALCS BELOW $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
    # Function for calculating k, Pcf and vapour flow state (for vapour phase)
    def Vapour_Conditions(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour':
            self.k_2.setEnabled(True)
            self.k_2.setReadOnly(False)
            self.pcf_2.setEnabled(True)
            self.pcf_2.setReadOnly(False)
            self.vapour_state_2.setEnabled(True)
            self.critical_flow_coefficient_2.setEnabled(True)
            self.critical_flow_coefficient_2.setReadOnly(False)
            self.subcritical_flow_coefficient_2.setEnabled(True)
            self.subcritical_flow_coefficient_2.setReadOnly(False)

            self.KN_2.setText('')
            self.KSH_2.setText('')
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            self.KSH_2.setStyleSheet('font: non-italic; color: black')
            self.KN_2.setEnabled(False)
            self.KSH_2.setEnabled(False)
            self.capacity_certification_requirement_2.setCurrentText('No')
            self.Kp_2.setText('')
            self.capacity_certification_requirement_2.setStyleSheet('font: non-italic; color: black')
            self.Kp_2.setStyleSheet('font: non-italic; color: black')
            self.capacity_certification_requirement.setEnabled(False)
            self.Kp_2.setEnabled(False)
            self.Kw_2.setText('')
            self.Kw_2.setEnabled(False)
            self.Re_2.setText('')
            self.Re_2.setEnabled(False)
            self.Kv_2.setText('')
            self.Kv_2.setEnabled(False)
            self.vol_flow_2.setText('')
            self.vol_flow_2.setEnabled(False)
        elif self.fluid_phase_2.currentText() != 'Vapour':
            self.k_2.setText('')
            self.k_2.setEnabled(False)
            self.k_2.setReadOnly(True)
            self.pcf_2.setText('')
            self.pcf_2.setEnabled(False)
            self.pcf_2.setReadOnly(True)
            self.vapour_state_2.setCurrentText('Edit manually here')
            self.vapour_state_2.setEnabled(False)
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            self.critical_flow_coefficient_2.setText('')
            self.critical_flow_coefficient_2.setEnabled(False)
            self.critical_flow_coefficient_2.setReadOnly(True)
            self.subcritical_flow_coefficient_2.setText('')
            self.subcritical_flow_coefficient_2.setEnabled(False)
            self.subcritical_flow_coefficient_2.setReadOnly(True)
            return None
        if self.specific_heat_cp_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            return None
        if self.specific_heat_cv_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            return None
        if self.relieving_pressure_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            return None
        if self.total_back_pressure_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            return None
        Cp = float(self.specific_heat_cp_2.text())
        Cv = float(self.specific_heat_cv_2.text())
        P1 = float(self.relieving_pressure_2.text()) + 1.01325
        P2 = float(self.total_back_pressure_2.text()) + 1.01325
        k = Cp/Cv
        Pcf = ((2/(k + 1))**(k/(k-1))) * P1
        flow_state = None
        k = round(k, 2)
        Pcf = round(Pcf, 2)
        self.k_2.setText(str(k))
        if self.k_2.text() == '':
            self.k_2.setText(str(k))
            self.k_2.setStyleSheet('font: italic; color: blue')
        elif self.k_2.text() == str(k):
            self.k_2.setStyleSheet('font: italic; color: blue')
        elif self.k_2.text() != str(k):
            self.k_2.setStyleSheet('font: non-italic; color: black')
        self.pcf_2.setText(str(Pcf))
        if self.pcf_2.text() == '':
            self.pcf_2.setText(str(Pcf))
            self.pcf_2.setStyleSheet('font: italic; color: blue')
        elif self.pcf_2.text() == str(Pcf):
            self.pcf_2.setStyleSheet('font: italic; color: blue')
        elif self.pcf_2.text() != str(Pcf):
            self.pcf_2.setStyleSheet('font: non-italic; color: black')
        if P2 < float(self.pcf_2.text()) or P2 == float(self.pcf_2.text()):
            self.vapour_state_2.setCurrentText('Critical')
            self.vapour_state_2.setStyleSheet('font: italic; color: blue')
            flow_state = 'Critical'
        elif P2 > float(self.pcf_2.text()):
            self.vapour_state_2.setCurrentText('Non-critical')
            self.vapour_state_2.setStyleSheet('font: italic; color: blue')
            flow_state = 'Non-critical'
        else:
            self.vapour_state_2.setCurrentText('Edit manually here')
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
        if self.vapour_state_2.currentText() != flow_state:
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
        elif self.vapour_state_2.currentText() == flow_state:
            self.vapour_state_2.setStyleSheet('font: italic; color: blue')
    def Vapour_Conditions_2(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour':
            self.k_2.setEnabled(True)
            self.k_2.setReadOnly(False)
            self.pcf_2.setEnabled(True)
            self.pcf_2.setReadOnly(False)
            self.vapour_state_2.setEnabled(True)
            self.critical_flow_coefficient_2.setEnabled(True)
            self.critical_flow_coefficient_2.setReadOnly(False)
            self.subcritical_flow_coefficient_2.setEnabled(True)
            self.subcritical_flow_coefficient_2.setReadOnly(False)
        elif self.fluid_phase_2.currentText() != 'Vapour':
            self.k_2.setText('')
            self.k_2.setEnabled(False)
            self.k_2.setReadOnly(True)
            self.pcf_2.setText('')
            self.pcf_2.setEnabled(False)
            self.pcf_2.setReadOnly(True)
            self.vapour_state_2.setCurrentText('Edit manually here')
            self.vapour_state_2.setEnabled(False)
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            self.critical_flow_coefficient_2.setText('')
            self.critical_flow_coefficient_2.setEnabled(False)
            self.critical_flow_coefficient_2.setReadOnly(True)
            self.subcritical_flow_coefficient_2.setText('')
            self.subcritical_flow_coefficient_2.setEnabled(False)
            self.subcritical_flow_coefficient_2.setReadOnly(True)
            return None
        if self.specific_heat_cp_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            return None
        if self.specific_heat_cv_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            return None
        if self.relieving_pressure_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            return None
        if self.total_back_pressure_2.text() == '':
            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            return None
        Cp = float(self.specific_heat_cp_2.text())
        Cv = float(self.specific_heat_cv_2.text())
        P1 = float(self.relieving_pressure_2.text()) + 1.01325
        P2 = float(self.total_back_pressure_2.text()) + 1.01325
        k = Cp/Cv
        Pcf = ((2/(k + 1))**(k/(k-1))) * P1
        flow_state = None
        k = round(k, 2)
        Pcf = round(Pcf, 2)
        if self.k_2.text() == '':
            self.k_2.setText(str(k))
            self.k_2.setStyleSheet('font: italic; color: blue')
        elif self.k_2.text() == str(k):
            self.k_2.setText(str(k))
            self.k_2.setStyleSheet('font: italic; color: blue')
        elif self.k_2.text() != str(k):
            self.k_2.setStyleSheet('font: non-italic; color: black')

        if self.pcf_2.text() == '':
            self.pcf_2.setText(str(Pcf))
            self.pcf_2.setStyleSheet('font: italic; color: blue')
        elif self.pcf_2.text() == str(Pcf):
            self.pcf_2.setText(str(Pcf))
            self.pcf_2.setStyleSheet('font: italic; color: blue')
        elif self.pcf_2.text() != str(Pcf):
            self.pcf_2.setStyleSheet('font: non-italic; color: black')
        if P2 < float(self.pcf_2.text()) or P2 == float(self.pcf_2.text()):
            flow_state = 'Critical'
        elif P2 > float(self.pcf_2.text()):
            flow_state = 'Non-critical'
        else:
            pass
        if self.vapour_state_2.currentText() == 'Edit manually here':
            self.vapour_state_2.setCurrentText(flow_state)
            self.vapour_state_2.setStyleSheet('font: italic; color: blue')
        elif self.vapour_state_2.currentText() == flow_state:
            self.vapour_state_2.setCurrentText(flow_state)
            self.vapour_state_2.setStyleSheet('font: italic; color: blue')
        elif self.vapour_state_2.currentText() != flow_state:
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')

    # Function for calculating coefficient flow (for vapour phase)
    def Vapour_Coefficient_Flow(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.specific_heat_cp_2.text() == '' or self.specific_heat_cv_2.text() == '':
            return None
        self.critical_flow_coefficient_2.setReadOnly(True)
        self.subcritical_flow_coefficient_2.setReadOnly(True)
        if self.vapour_state_2.currentText() == 'Critical':
            k = float(self.specific_heat_cp_2.text()) / float(self.specific_heat_cv_2.text())
            C = 0.03948 * ((k*(2/(k + 1))**((k+1) / (k-1)))**(0.5))
            C = round(C, 5)
            self.critical_flow_coefficient_2.setText(str(C))
            self.critical_flow_coefficient_2.setStyleSheet('font: italic; color: blue')
            self.subcritical_flow_coefficient_2.setText('NA')
            self.subcritical_flow_coefficient_2.setStyleSheet('font: italic; color: blue')
        elif self.vapour_state_2.currentText() == 'Non-critical':
            k = float(self.specific_heat_cp_2.text()) / float(self.specific_heat_cv_2.text())
            r = (float(self.total_back_pressure_2.text()) + 1.01325) / (float(self.relieving_pressure_2.text()) + 1.01325)
            F2 = ((k/(k-1))*(r**(2/k))*((1-r**((k-1)/k))/(1-r)))**0.5
            F2 = round(F2, 5)
            self.critical_flow_coefficient_2.setText('NA')
            self.critical_flow_coefficient_2.setStyleSheet('font: italic; color: blue')
            self.subcritical_flow_coefficient_2.setText(str(F2))
            self.subcritical_flow_coefficient_2.setStyleSheet('font: italic; color: blue')

    # Function for calculating the variable Kd (for vapour phase)
    def Vapour_Kd_Calc(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Critical':
            pass
        elif self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Non-critical':
            pass
        else:
            self.Kd_2.setText('')
            self.Kd_2.setStyleSheet('font: non-italic; color: black')
            return None
        Kd = None
        if self.rupture_disk_2.currentText() == 'No':
            Kd = 0.975
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kd = 0.62
        self.Kd_2.setText(str(Kd))
        self.Kd_2.setStyleSheet('font: italic; color: blue')
    def Vapour_Kd_Calc_2(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Critical':
            pass
        elif self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Non-critical':
            pass
        else:
            self.Kd_2.setText('')
            self.Kd_2.setStyleSheet('font: non-italic; color: black')
            return None
        Kd = None
        if self.rupture_disk_2.currentText() == 'No':
            Kd = 0.975
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kd = 0.62
        if self.Kd_2.text() == '':
            self.Kd_2.setText(str(Kd))
            self.Kd_2.setStyleSheet('font: italic; color: blue')
        elif self.Kd_2.text() == str(Kd):
            self.Kd_2.setText(str(Kd))
            self.Kd_2.setStyleSheet('font: italic; color: blue')
        elif self.Kd_2.text() != str(Kd):
            self.Kd_2.setStyleSheet('font: non-italic; color: black')

    # Function for calculating the variable Kb (for vapour phase)
    def Vapour_Kb_Calc(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Critical':
            self.Kb_2.setEnabled(True)
            self.Kb_2.setReadOnly(False)
            pass
        elif self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Non-critical':
            self.Kb_2.setText('NA')
            self.Kb_2.setReadOnly(True)
            self.Kb_2.setEnabled(False)
            self.Kb_2.setStyleSheet('font: italic; color: blue')
            return None
        else:
            self.Kb_2.setText('')
            self.Kb_2.setStyleSheet('font: non-italic; color: black')
            return None
        Kb = None
        if self.design_type_2.currentText() == 'Conventional':
            Kb = 1
            self.Kb_2.setText(str(Kb))
            self.Kb_2.setStyleSheet('font: italic; color: blue')
        elif self.design_type_2.currentText() != 'Conventional':
            self.Kb_2.setText('Input manually (non-conventional valve)')
            self.Kb_2.setStyleSheet('font: italic; color: blue')
    def Vapour_Kb_Calc_2(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Critical':
            self.Kb_2.setEnabled(True)
            self.Kb_2.setReadOnly(False)
            pass
        elif self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Non-critical':
            self.Kb_2.setText('NA')
            self.Kb_2.setReadOnly(True)
            self.Kb_2.setEnabled(False)
            self.Kb_2.setStyleSheet('font: italic; color: blue')
        else:
            self.Kb_2.setText('')
            self.Kb_2.setStyleSheet('font: non-italic; color: black')
            return None
        Kb = None
        if self.design_type_2.currentText() == 'Conventional':
            Kb = 1.0
        elif self.design_type_2.currentText() != 'Conventional':
            Kb = str('Must input manually (non conventional valve)')

        if self.Kb_2.text() == '':
            self.Kb_2.setText(str(Kb))
            self.Kb_2.setStyleSheet('font: italic; color: blue')
        elif self.Kb_2.text() == str(Kb):
            self.Kb_2.setText(str(Kb))
            self.Kb_2.setStyleSheet('font: italic; color: blue')
        elif self.Kb_2.text() != str(Kb):
            self.Kb_2.setStyleSheet('font: non-italic; color: black')

    # Function for calculating variable Kc (for vapour phase)
    def Vapour_Kc_Calc(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass

        if self.rupture_disk_2.currentText() == 'No':
            Kc = 1.0
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kc = 0.9

        self.Kc_2.setText(str(Kc))
        self.Kc_2.setStyleSheet('font: italic; color: blue')
    def Vapour_Kc_Calc_2(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Critical':
            pass
        elif self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Non-critical':
            pass
        else:
            return None

        if self.rupture_disk_2.currentText() == 'No':
            Kc = 1.0
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kc = 0.9

        if self.Kc_2.text() == '':
            self.Kc_2.setText(str(Kc))
            self.Kc_2.setStyleSheet('font: italic; color: blue')
        elif self.Kc_2.text() == str(Kc):
            self.Kc_2.setStyleSheet('font: italic; color: blue')
        elif self.Kc_2.text() != str(Kc):
            self.Kc_2.setStyleSheet('font: non-italic; color: black')

    # Function for when you have balanced bellows (for vapour phase)
    def Balanced_PRV(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        if self.design_type_2.currentText() == 'Conventional' and self.fluid_phase_2.currentText() == 'Vapour':
            return None
        elif 'Balanced' in self.design_type_2.currentText() and self.fluid_phase_2.currentText() == 'Vapour':
            self.vapour_state_2.setCurrentText('Critical')
            self.vapour_state_2.setStyleSheet('font: italic; color: blue')

    # Function for calculating variable A (for vapour phase)
    def Vapour_Area_Calc(self):
        if (self.component_1_nist_identification.text() == 'Water' and self.total_composition_2.text() == str(1.0)
        and float(self.relieving_temp_2.text()) < 649):
            return None
        else:
            pass
        def Is_Number(number):
            try:
                float(number)
                return True
            except ValueError:
                return False
        self.discharge_area_2.setReadOnly(True)
        self.discharge_area_2.setEnabled(True)
        if self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Critical':
            if (Is_Number(self.mass_flow_2.text()) == True and
                Is_Number(self.critical_flow_coefficient_2.text()) == True and
                Is_Number(self.Kd_2.text()) == True and
                Is_Number(self.relieving_pressure_2.text()) == True and
                Is_Number(self.Kb_2.text()) == True and
                Is_Number(self.Kc_2.text()) == True and
                Is_Number(self.relieving_temp_2.text()) == True and
                Is_Number(self.compressibility_factor_2.text()) == True and
                Is_Number(self.molecular_weight_2.text()) == True):
                W = float(self.mass_flow_2.text())
                C = float(self.critical_flow_coefficient_2.text())
                Kd = float(self.Kd_2.text())
                P1 = (float(self.relieving_pressure_2.text()) + 1.01325)*100
                Kb = float(self.Kb_2.text())
                Kc = float(self.Kc_2.text())
                T = float(self.relieving_temp_2.text()) + 273
                Z = float(self.compressibility_factor_2.text())
                M = float(self.molecular_weight_2.text())
                A = (W / (C*Kd*P1*Kb*Kc)) * ((T*Z)/M)**(0.5)
                A = round(A,2)
                self.discharge_area_2.setText(str(A))
                self.discharge_area_2.setStyleSheet('font: italic; color: blue')
            else:
                self.discharge_area_2.setText('')
                return None
        elif self.fluid_phase_2.currentText() == 'Vapour' and self.vapour_state_2.currentText() == 'Non-critical':
            if (Is_Number(self.mass_flow_2.text()) == True and
                Is_Number(self.subcritical_flow_coefficient_2.text()) == True and
                Is_Number(self.Kd_2.text()) == True and
                Is_Number(self.Kc_2.text()) == True and
                Is_Number(self.relieving_temp_2.text()) == True and
                Is_Number(self.compressibility_factor_2.text()) == True and
                Is_Number(self.molecular_weight_2.text()) == True and
                Is_Number(self.relieving_pressure_2.text()) == True and
                Is_Number(self.total_back_pressure_2.text()) == True):
                W = float(self.mass_flow_2.text())
                F2 = float(self.subcritical_flow_coefficient_2.text())
                Kd = float(self.Kd_2.text())
                Kc = float(self.Kc_2.text())
                T = float(self.relieving_temp_2.text()) + 273
                Z = float(self.compressibility_factor_2.text())
                M = float(self.molecular_weight_2.text())
                P1 = (float(self.relieving_pressure_2.text()) + 1.01325)*100
                P2 = (float(self.total_back_pressure_2.text()) + 1.01325)*100
                A = ((17.9*W)/(F2*Kd*Kc)) * (((T*Z)/(M*P1*(P1-P2)))**(0.5))
                A = round(A,2)
                self.discharge_area_2.setText(str(A))
                self.discharge_area_2.setStyleSheet('font: italic; color: blue')
            else:
                self.discharge_area_2.setText('')
                return None
        else:
            self.discharge_area_2.setText('')
            return None
    #VAPOUR CALCS ABOVE $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

    #STEAM CALCS BELOW $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
    # Functions for calculating steam conditions KN and KSH
    def Steam_Conditions(self):
        if (self.component_1_nist_identification.text() == 'Water' and
            self.total_composition_2.text() == str(1.0) and
            self.fluid_phase_2.currentText() == 'Vapour' and
            float(self.relieving_temp_2.text()) < 649):
            self.KN_2.setEnabled(True)
            self.KN_2.setReadOnly(False)
            self.KSH_2.setEnabled(True)
            self.KSH_2.setReadOnly(False)
            self.Kb_2.setEnabled(True)

            self.k_2.setText('')
            self.pcf_2.setText('')
            self.vapour_state_2.setCurrentText('Edit manually here')
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            self.critical_flow_coefficient_2.setText('')
            self.subcritical_flow_coefficient_2.setText('')
            self.k_2.setEnabled(False)
            self.pcf_2.setEnabled(False)
            self.vapour_state_2.setEnabled(False)
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            self.critical_flow_coefficient_2.setEnabled(False)
            self.subcritical_flow_coefficient_2.setEnabled(False)
            self.capacity_certification_requirement_2.setCurrentText('No')
            self.Kp_2.setText('')
            self.capacity_certification_requirement_2.setEnabled(False)
            self.Kp_2.setEnabled(False)
            self.Kw_2.setText('')
            self.Kw_2.setEnabled(False)
            self.Re_2.setText('')
            self.Re_2.setEnabled(False)
            self.Kv_2.setText('')
            self.Kv_2.setEnabled(False)
            self.vol_flow_2.setText('')
            self.vol_flow_2.setEnabled(False)
        else:
            self.KN_2.setText('')
            self.KSH_2.setText('')
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            self.KSH.setStyleSheet('font: non-italic; color: black')
            self.KN_2.setEnabled(False)
            self.KN_2.setReadOnly(True)
            self.KSH_2.setEnabled(False)
            self.KSH_2.setReadOnly(True)
            return None
        if self.relieving_pressure_2.text() == '':
            self.KN_2.setText('')
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            return None
        elif self.set_pressure_2.text() == '':
            self.KSH_2.setText('')
            self.KSH_2.setStyleSheet('font: non-italic; color: black')
            return None
        else:
            pass
        P1 = (float(self.relieving_pressure_2.text()) + 1.01325)*100
        if P1 < 10339:
            KN = 1.0
        elif P1 > 10339 and P1 < 22057:
            KN = (0.02764*P1 - 1000) / (0.03324*P1 - 1061)
            KN = round(KN, 2)
        elif P1 > 22057:
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            return None
        self.KN_2.setText(str(KN))
        self.KN_2.setStyleSheet('font: italic; color:blue')

        df = pd.read_excel('KSH Values.xlsx')
        P_Set = float(self.set_pressure_2.text())*100
        T = float(self.relieving_temp_2.text())

        if T < 149 or T > 649 or P_Set < 103 or P_Set > 20679:
            self.KSH_2.setText('No value found, please input manually')
            self.KSH_2.setStyleSheet('font: italic; color: blue')
        else:
            pass

        if self.KSH_2.text() == 'No value found, please input manually':
            KSH = None
        else:
            for i in range(1, 10):
                if T <= df.columns[i + 1] and T >= df.columns[i]:
                    x1 = df.columns[i]
                    x2 = df.columns[i + 1]

            def Temp_Interp(y1, y2):
                if y1 == '-' or y2 == '-':
                    return '-'
                else:
                    y = y1 + (T - x1)*((y2 - y1)/(x2 - x1))
                    y = round(y ,2)
                    return y
            df[T] = df.apply(lambda x: Temp_Interp(x[x1], x[x2]), axis = 1)

            for i in range(27):
                if df[T].iloc[i + 1] == '-':
                    KSH = None
                elif P_Set <= df['Set Pressure (kPag)'].iloc[i + 1] and P_Set >= df['Set Pressure (kPag)'].iloc[i]:
                    y1 = df[T].iloc[i]
                    y2 = df[T].iloc[i + 1]
                    x1 = df['Set Pressure (kPag)'].iloc[i]
                    x2 = df['Set Pressure (kPag)'].iloc[i + 1]

                    y = y1 + (P_Set - x1)*((y2 - y1)/(x2 - x1))
                    y = round(y, 2)
                    KSH = y
                    break
        if KSH == None:
            self.KSH_2.setText('No value found, please input manually')
            self.KSH_2.setStyleSheet('font: italic; color: blue')
        else:
            self.KSH_2.setText(str(KSH))
            self.KSH_2.setStyleSheet('font: italic; color: blue')
    def Steam_Conditions_2(self):
        if (self.component_1_nist_identification.text() == 'Water' and
            self.total_composition_2.text() == str(1.0) and
            self.fluid_phase_2.currentText() == 'Vapour' and
            float(self.relieving_temp_2.text()) < 649):
            self.KN_2.setEnabled(True)
            self.KN_2.setReadOnly(False)
            self.KSH_2.setEnabled(True)
            self.KSH_2.setReadOnly(False)
        else:
            self.KN_2.setText('')
            self.KSH_2.setText('')
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            self.KSH.setStyleSheet('font: non-italic; color: black')
            self.KN_2.setEnabled(False)
            self.KN_2.setReadOnly(True)
            self.KSH_2.setEnabled(False)
            self.KSH_2.setReadOnly(True)
        if self.relieving_pressure_2.text() == '':
            self.KN_2.setText('')
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            return None
        elif self.set_pressure_2.text() == '':
            self.KSH_2.setText('')
            self.KSH_2.setStyleSheet('font: non-italic; color: black')
            return None
        else:
            pass
        P1 = (float(self.relieving_pressure_2.text()) + 1.01325)*100
        if P1 < 10339:
            KN = 1.0
        elif P1 > 10339 and P1 < 22057:
            KN = (0.02764*P1 - 1000) / (0.03324*P1 - 1061)
            KN = round(KN, 2)
        elif P1 > 22057:
            self.KN_2.setStyleSheet('font: non-italic; color: black')
            return None
        if self.KN_2.text() == '':
            self.KN_2.setText(str(KN))
            self.KN_2.setStyleSheet('font: italic; color:blue')
        elif self.KN_2.text() == str(KN):
            self.KN_2.setStyleSheet('font: italic; color:blue')
        elif self.KN_2.text() != str(KN):
            self.KN_2.setStyleSheet('font: non-italic; color: black')

        df = pd.read_excel('KSH Values.xlsx')
        P_Set = float(self.set_pressure_2.text())*100
        T = float(self.relieving_temp_2.text())

        KSH = 0
        if T < 149 or T > 649 or P_Set < 103 or P_Set > 20679:
            KSH = None
        else:
            pass

        if KSH == None:
            pass
        else:
            for i in range(1, 10):
                if T <= df.columns[i + 1] and T >= df.columns[i]:
                    x1 = df.columns[i]
                    x2 = df.columns[i + 1]

            def Temp_Interp(y1, y2):
                if y1 == '-' or y2 == '-':
                    return '-'
                else:
                    y = y1 + (T - x1)*((y2 - y1)/(x2 - x1))
                    y = round(y ,2)
                    return y
            df[T] = df.apply(lambda x: Temp_Interp(x[x1], x[x2]), axis = 1)

            for i in range(27):
                if df[T].iloc[i + 1] == '-':
                    KSH = None
                elif P_Set <= df['Set Pressure (kPag)'].iloc[i + 1] and P_Set >= df['Set Pressure (kPag)'].iloc[i]:
                    y1 = df[T].iloc[i]
                    y2 = df[T].iloc[i + 1]
                    x1 = df['Set Pressure (kPag)'].iloc[i]
                    x2 = df['Set Pressure (kPag)'].iloc[i + 1]

                    y = y1 + (P_Set - x1)*((y2 - y1)/(x2 - x1))
                    y = round(y, 2)
                    KSH = y
                    break
        if KSH == None:
            if self.KSH_2.text() == '':
                self.KSH_2.setText('No value found, please input manually')
                self.KSH_2.setStyleSheet('font: italic; color: blue')
            elif self.KSH_2.text() == 'No value found, please input manually':
                self.KSH_2.setStyleSheet('font: italic; color: blue')
            elif self.KSH_2.text() != 'No value found, please input manually':
                self.KSH_2.setStyleSheet('font: non-italic; color: black')
        else:
            if self.KSH_2.text() == '':
                self.KSH_2.setText(str(KSH))
                self.KSH_2.setStyleSheet('font: italic; color: blue')
            elif self.KSH_2.text() == str(KSH):
                self.KSH_2.setStyleSheet('font: italic; color: blue')
            elif self.KSH_2.text() != str(KSH):
                self.KSH_2.setStyleSheet('font: non-italic; color: black')

    # Functions for calculating variable Kd for steam conditions
    def Steam_Kd_Calc(self):
        if (self.component_1_nist_identification.text() == 'Water' and
            self.total_composition_2.text() == str(1.0) and
            self.fluid_phase_2.currentText() == 'Vapour' and
            float(self.relieving_temp_2.text()) < 649):
            pass
        else:
            return None
        if self.rupture_disk_2.currentText() == 'No':
            Kd = 0.975
            self.Kd_2.setText(str(Kd))
            self.Kd_2.setStyleSheet('font: italic; color: blue')
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kd = 0.62
            self.Kd_2.setText(str(Kd))
            self.Kd_2.setStyleSheet('font: italic; color: blue')
    def Steam_Kd_Calc_2(self):
        if (self.component_1_nist_identification.text() == 'Water' and
            self.total_composition_2.text() == str(1.0) and
            self.fluid_phase_2.currentText() == 'Vapour' and
            float(self.relieving_temp_2.text()) < 649):
            pass
        else:
            return None
        Kd = None
        if self.rupture_disk_2.currentText() == 'No':
            Kd = 0.975
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kd = 0.62

        if self.Kd_2.text() == '':
            self.Kd_2.setText(str(Kd))
            self.Kd_2.setStyleSheet('font: italic; color: blue')
        elif self.Kd_2.text() == str(Kd):
            self.Kd_2.setStyleSheet('font: italic; color: blue')
        elif self.Kd_2.text() != str(Kd):
            self.Kd_2.setStyleSheet('font: non-italic; color: black')

    # Functions for calculating variable Kb for steam conditions
    def Steam_Kb_Calc(self):
        if self.design_type_2.currentText() == 'Conventional':
            self.Kb_2.setText(str(1.0))
            self.Kb_2.setStyleSheet('font: italic; color: blue')
        elif self.design_type_2.currentText() != 'Conventional':
            self.Kb_2.setText('')
            self.Kb_2.setStyleSheet('font: non-italic; color: black')
    def Steam_Kb_Calc_2(self):
        if self.design_type_2.currentText() == 'Conventional':
            if self.Kb_2.text() == '':
                self.Kb_2.setText(str(1.0))
                self.Kb_2.setStyleSheet('font: italic; color: blue')
            elif self.Kb_2.text() == str(1.0):
                self.Kb_2.setText(str(1.0))
                self.Kb_2.setStyleSheet('font: italic; color: blue')
            elif self.Kb_2.text() != str(1.0):
                self.Kb_2.setStyleSheet('font: non-italic; color: black')
        elif self.design_type_2.currentText() != 'Conventional':
            self.Kb_2.setText('')
            self.Kb_2.setStyleSheet('font: non-italic; color: black')

    # Functions for calculating variable
    def Steam_Kc_Calc(self):
        if self.rupture_disk_2.currentText() == 'No':
            self.Kc_2.setText(str(1.0))
            self.Kc_2.setStyleSheet('font: italic; color: blue')
        elif self.rupture_disk_2.currentText() == 'Yes':
            self.Kc_2.setText(str(0.9))
            self.Kc_2.setStyleSheet('font: italic; color: blue')
    def Steam_Kc_Calc_2(self):
        if self.rupture_disk_2.currentText() == 'No':
            Kc = 1.0
        elif self.rupture_disk_2.currentText() == 'Yes':
            Kc = 0.9

        if self.Kc_2.text() == '':
            self.Kc_2.setText(str(Kc))
            self.Kc_2.setStyleSheet('font: italic; color: blue')
        elif self.Kc_2.text() == str(Kc):
            self.Kc_2.setStyleSheet('font: italic; color: blue')
        elif self.Kc_2.text() != str(Kc):
            self.Kc_2.setStyleSheet('font: non-italic; color: black')

    # Function for calculating variable A (for steam conditions)
    def Steam_Area_Calc(self):
        if (self.component_1_nist_identification.text() == 'Water' and
            self.total_composition_2.text() == str(1.0) and
            self.fluid_phase_2.currentText() == 'Vapour' and
            float(self.relieving_temp_2.text()) < 649):
            pass
        else:
            return None
        def Is_Number(number):
            try:
                float(number)
                return True
            except ValueError:
                return False
        self.discharge_area_2.setReadOnly(True)
        self.discharge_area_2.setEnabled(True)

        if self.mass_flow_2.text() == '' or self.relieving_pressure_2.text() == '':
            return None
        if (Is_Number(float(self.mass_flow_2.text())) == True and
            Is_Number(float(self.relieving_pressure_2.text())) == True and
            Is_Number(float(self.Kd_2.text())) == True and
            Is_Number(float(self.Kb_2.text())) == True and
            Is_Number(float(self.Kc_2.text())) == True and
            Is_Number(float(self.KN_2.text())) == True and
            Is_Number(float(self.KSH_2.text())) == True):

            W = float(self.mass_flow_2.text())
            P1 = (float(self.relieving_pressure_2.text()) + 1.01325)*100
            Kd = float(self.Kd_2.text())
            Kb = float(self.Kb_2.text())
            Kc = float(self.Kc_2.text())
            KN = float(self.KN_2.text())



            KSH = float(self.KSH_2.text())

            A = (190.5*W)/(P1*Kd*Kb*Kc*KN*KSH)
            A = round(A, 2)

            self.discharge_area_2.setText(str(A))
            self.discharge_area_2.setStyleSheet('font: italic; color: blue')

        else:
            self.discharge_area_2.setText('')
            return None
    #STEAM CALCS ABOVE $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

    #LIQUID CALCS BELOW $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
    # Function for activating liquid conditions (for liquid conditions with capacity certification)
    def Liquid_Activation(self):
        if self.fluid_phase_2.currentText() == 'Liquid':
            self.capacity_certification_requirement_2.setEnabled(True)
            self.Kp_2.setEnabled(True)
            self.Kw_2.setEnabled(True)
            self.Re_2.setEnabled(True)
            self.Kv_2.setEnabled(True)
            self.vol_flow_2.setEnabled(True)
            self.vol_flow_2.setReadOnly(True)
            self.discharge_area_2.setReadOnly(True)

            self.k_2.setText('')
            self.k_2.setEnabled(False)
            self.pcf_2.setText('')
            self.pcf_2.setEnabled(False)
            self.vapour_state_2.setEnabled(False)
            self.vapour_state_2.setStyleSheet('font: non-italic; color: black')
            self.critical_flow_coefficient_2.setText('')
            self.critical_flow_coefficient_2.setEnabled(False)
            self.subcritical_flow_coefficient_2.setText('')
            self.subcritical_flow_coefficient_2.setEnabled(False)
            self.KN_2.setText('')
            self.KN_2.setEnabled(False)
            self.KSH_2.setText('')
            self.KSH_2.setEnabled(False)
            #self.Kd_2.setText('')
            #self.Kd_2.setStyleSheet('font: non-italic; color: black')
            self.Kb_2.setText('')
            self.Kb_2.setEnabled(False)
            self.Kb_2.setStyleSheet('font: non-italic; color: black')
            self.discharge_area_2.setText('')

    # Function for calculating Kw, Kd and Kc (for liquid conditions with capacity certification)
    def Capacity_Certification_Yes(self):
        if (self.fluid_phase_2.currentText() == 'Liquid' and
            self.capacity_certification_requirement_2.currentText() == 'Yes'):
            pass
        else:
            return None

        self.Kp_2.setText('')
        self.Kp_2.setEnabled(False)

        if self.total_back_pressure_2.text() == '' or self.set_pressure_2.text() == '':
            pass
        else:
            if float(self.total_back_pressure_2.text()) == 0:
                Kw = 1.0
            elif self.design_type_2.currentText() != 'Conventional':
                PB = (float(self.total_back_pressure_2.text()) + 1.0325)*14.5038 - 14.6959
                PS = (float(self.set_pressure_2.text()) + 1.01325)*14.5038 - 14.6959
                percent_backpressure = (PB/PS)*100
                if percent_backpressure <= 15:
                    Kw = 1.0
                elif percent_backpressure > 15 and percent_backpressure <= 20:
                    Kw = -0.001*(percent_backpressure**2) + 0.0312*percent_backpressure + 0.7465
                    Kw = round(Kw, 3)
                elif percent_backpressure > 20 and percent_backpressure <= 50:
                    Kw = -0.01*percent_backpressure + 1.175
                    Kw = round(Kw, 3)
                elif percent_backpressure > 50:
                    Kw = 'No value found, input manually'
            elif self.design_type_2.currentText() == 'Conventional':
                Kw = 'No value found, input manually'

            self.Kw_2.setText(str(Kw))
            self.Kw_2.setStyleSheet('font: italic; color: blue')

            if self.rupture_disk_2.currentText() == 'No':
                Kd = 0.65
                self.Kd_2.setText(str(Kd))
                self.Kd_2.setStyleSheet('font: italic; color : blue')
            elif self.rupture_disk_2.currentText() == 'Yes':
                Kd = 0.62
                self.Kd_2.setText(str(Kd))
                self.Kd_2.setStyleSheet('font: italic; color : blue')

            if self.rupture_disk_2.currentText() == 'No':
                Kc = 1.0
            elif self.rupture_disk_2.currentText() == 'Yes':
                Kc = 0.9

            self.Kc_2.setText(str(Kc))
            self.Kc_2.setStyleSheet('font: italic; color: blue')

            self.Re_2.setText('Input missing for calculation')
            self.Re_2.setStyleSheet('font: italic; color: blue')
            self.Re_2.setReadOnly(True)
            self.Kv_2.setText('Input missing for calculation')
            self.Kv_2.setStyleSheet('font: italic; color: blue')
            self.Kv_2.setReadOnly(True)
    def Capacity_Certification_Yes_2(self):
        if (self.fluid_phase_2.currentText() == 'Liquid' and
            self.capacity_certification_requirement_2.currentText() == 'Yes'):
            pass
        else:
            return None
        self.Kp_2.setText('')
        self.Kp_2.setEnabled(False)
        if self.total_back_pressure_2.text() == '' or self.set_pressure_2.text() == '':
            pass
        else:
            if float(self.total_back_pressure_2.text()) == 0:
                Kw = 1.0
            elif self.design_type_2.currentText() != 'Conventional':
                PB = (float(self.total_back_pressure_2.text()) + 1.0325)*14.5038 - 14.6959
                PS = (float(self.set_pressure_2.text()) + 1.01325)*14.5038 - 14.6959
                percent_backpressure = (PB/PS)*100
                if percent_backpressure <= 15:
                    Kw = 1.0
                elif percent_backpressure > 15 and percent_backpressure <= 20:
                    Kw = -0.001*(percent_backpressure**2) + 0.0312*percent_backpressure + 0.7465
                    Kw = round(Kw, 3)
                elif percent_backpressure > 20 and percent_backpressure <= 50:
                    Kw = -0.01*percent_backpressure + 1.175
                    Kw = round(Kw, 3)
                elif percent_backpressure > 50:
                    Kw = 'No value found, input manually'
            elif self.design_type_2.currentText() == 'Conventional':
                Kw = 'No value found, input manually'

            if self.Kw_2.text() == '':
                self.Kw_2.setText(str(Kw))
                self.Kw_2.setStyleSheet('font: italic; color: blue')
            elif self.Kw_2.text() == str(Kw):
                self.Kw_2.setStyleSheet('font: italic; color: blue')
            elif self.Kw_2.text() != str(Kw):
                self.Kw_2.setStyleSheet('font: non-italic; color: black')

            if self.rupture_disk_2.currentText() == 'No':
                Kd = 0.65
            elif self.rupture_disk_2.currentText() == 'Yes':
                Kd = 0.62

            if self.Kd_2.text() == '':
                self.Kd_2.setText(str(Kd))
                self.Kd_2.setStyleSheet('font: italic; color: blue')
            elif self.Kd_2.text() == str(Kd):
                self.Kd_2.setStyleSheet('font: italic; color: blue')
            elif self.Kd_2.text() != str(Kd):
                self.Kd_2.setStyleSheet('font: non-italic; color: black')

            if self.rupture_disk_2.currentText() == 'No':
                Kc = 1.0
            elif self.rupture_disk_2.currentText() == 'Yes':
                Kc = 0.9

            if self.Kc_2.text() == '':
                self.Kc_2.setText(str(Kc))
                self.Kc_2.setStyleSheet('font: italic; color: blue')
            elif self.Kc_2.text() == str(Kc):
                self.Kc_2.setStyleSheet('font: italic; color: blue')
            elif self.Kc_2.text() != str(Kc):
                self.Kc_2.setStyleSheet('font: non-italic; color: black')

    # Function for calculating volumetric flow from mass flow (for liquid conditions with capacity certification)
    def Vol_Flow_Calc(self):
        if (self.fluid_phase_2.currentText() == 'Liquid' and
            self.capacity_certification_requirement_2.currentText() == 'Yes'):
            pass
        else:
            return None
        self.Kp_2.setText('')
        self.Kp_2.setEnabled(False)
        self.vol_flow_2.setReadOnly(True)
        if self.mass_flow_2.text() == '' or self.density_2.text() == '':
            self.vol_flow_2.setText('')
            self.vol_flow_2.setStyleSheet('font: non-italic; color: black')
            return None
        else:
            pass
        M = float(self.mass_flow_2.text())
        D = float(self.density_2.text())
        Vol = (M/D)*16.6667
        Vol = round(Vol, 2)
        self.vol_flow_2.setText(str(Vol))
        self.vol_flow_2.setStyleSheet('font: italic; color: blue')

    # Function for calculating Area, Re and Kv
    def Liquid_Area_Calc_Capacity_Cert(self):
        if (self.fluid_phase_2.currentText() == 'Liquid' and
            self.capacity_certification_requirement_2.currentText() == 'Yes'):
            pass
        else:
            return None
        def Is_Number(number):
            try:
                float(number)
                return True
            except ValueError:
                return False
        if (Is_Number(self.vol_flow_2.text()) == False or
            Is_Number(self.Kd_2.text()) == False or
            Is_Number(self.Kw_2.text()) == False or
            Is_Number(self.Kc_2.text()) == False or
            Is_Number(self.density_2.text()) == False or
            Is_Number(self.relieving_pressure_2.text()) == False or
            Is_Number(self.total_back_pressure_2.text()) == False or
            Is_Number(self.viscosity_2.text()) == False or
            self.capacity_certification_requirement_2.currentText() == 'No'):
            return None
        else:
            pass
        Q = float(self.vol_flow_2.text())
        Kd = float(self.Kd_2.text())
        Kw = float(self.Kw_2.text())
        Kc = float(self.Kc_2.text())
        Gl = float(self.density_2.text()) / 1000
        P1 = (float(self.relieving_pressure_2.text()) + 1.01325)*100 - 101.325
        P2 = (float(self.total_back_pressure_2.text()) + 1.01325)*100 - 101.325
        VIS = float(self.viscosity_2.text())

        def f(A):
            return A - ((11.78*Q)/(Kd*Kw*Kc*(1/(0.9935 + 2.878/((18800*Q*Gl)/(VIS*(A**0.5))**0.5) + 342.75/((18800*Q*Gl)/(VIS*(A**0.5))**1.5)))))*((Gl/(P1-P2))**0.5)

        A = fsolve(f, 1)[0]

        self.discharge_area_2.setText(str(A))
        self.discharge_area_2.setStyleSheet('font: italic; color: blue')

        Re = (Q*18800*Gl)/(VIS*(A**0.5))
        Kv = 1/(0.9935 + 2.878/(Re**0.5) + 342.75/(Re**1.5))

        A = round(A, 2)
        Kv = round(Kv, 2)
        Re = f"{Decimal(str(Re)):.2E}"

        self.discharge_area_2.setText(str(A))
        self.discharge_area_2.setStyleSheet('font: italic; color: blue')
        self.Kv_2.setText(str(Kv))
        self.Kv_2.setStyleSheet('font: italic; color: blue')
        self.Re_2.setText(str(Re))
        self.Re_2.setStyleSheet('font: italic; color: blue')
    def Liquid_Area_Calc_Capacity_Cert_2(self):
        if (self.fluid_phase_2.currentText() == 'Liquid' and
            self.capacity_certification_requirement_2.currentText() == 'Yes'):
            pass
        else:
            return None
        if (self.mass_flow_2.text() == '' or
            self.Kw_2.text() == '' or
            self.density_2.text() == '' or
            self.viscosity_2.text() == ''):
            self.Re_2.setText('Input missing for calculation')
            self.Re_2.setStyleSheet('font: italic; color: blue')
            self.Kv_2.setText('Input missing for calculation')
            self.Kv_2.setStyleSheet('font: italic; color: blue')
        else:
            pass
# End of datasheet and results section ---------------------------------------------------------------------------------------------------------------



# Start of code for pressing 'generate datasheet' button ---------------------------------------------------------------------------------------------
    def Document_Information(self):
        ws['C1'].value = ws['C1'].value + '     ' + self.sheet_number_2.text()
        ws['C2'].value = ws['C2'].value + '     ' + self.requisition_number_2.text()
        ws['C3'].value = ws['C3'].value + '     ' + self.job_number_2.text()
        ws['C4'].value = ws['C4'].value + '     ' + self.date_2.text()
        ws['C5'].value = ws['C5'].value + '     ' + self.revision_2.text()
        ws['C6'].value = ws['C6'].value + '     ' + self.by_2.text()
        wb.save('Datasheets/Datasheet.xlsx')

    def General_Equipment_Information(self):
        ws['B8'].value = ws['B8'].value + '     ' + self.item_number_2.text()
        ws['B9'].value = ws['B9'].value + '     ' + self.tag_number_2.text()
        ws['B10'].value = ws['B10'].value + '     ' + self.service_line_number_2.text()
        ws['B11'].value = ws['B11'].value + '     ' + self.number_required_2.text()
        wb.save('Datasheets/Datasheet.xlsx')

    def Selection_Basis(self):
        if self.asme_code_2.currentText() == 'Yes':
            ws['D8'].value = 'Code: ASME VIII'
        elif self.asme_code_2.currentText() == 'Yes with U stamp':
            ws['D8'].value = 'Code: ASME VIII with U stamp'
        elif self.asme_code_2.currentText() == 'No':
            ws['D8'].value = 'Code: ' + self.asme_code_3.text()

        if self.api_code_2.currentText() == 'Yes':
            ws['D9'].value = 'Comply with API 526: ' + self.api_code_2.currentText()
        elif self.api_code_2.currentText() == 'No':
            ws['D9'].value = 'Comply with code: ' + self.api_code_3.text()

        ws['D10'].value = 'Fire Conditions: ' + self.fire_condition_2.currentText()
        ws['D11'].value = 'Rupture Disk: ' + self.rupture_disk_2.currentText()

        wb.save('Datasheets/Datasheet.xlsx')

    def Valve_Design(self):
        ws['B13'].value = 'Design Type: ' + self.design_type_2.currentText()
        if self.nozzle_type_2.currentText() == 'Other':
            ws['B15'].value = 'Design Type: ' + self.nozzle_type_3.text()
        elif self.nozzle_type_2.currentText() != 'Other':
            ws['B15'].value = 'Design Type: ' + self.nozzle_type_2.currentText()

        ws['B17'].value = 'Bonnet Type: ' + self.bonnet_type_2.currentText()
        if self.seat_tightness_2.currentText() == 'API 527':
            ws['B19'].value = 'Seat Tightness: ' + self.seat_tightness_2.currentText()
        elif self.seat_tightness_2.currentText() == 'Other':
            ws['B19'].value = 'Seat Tightness: ' + self.seat_tightness_3.text()
        wb.save('Datasheets/Datasheet.xlsx')

    def Connections(self):
        if self.inlet_facing_2.currentText() == 'Raised Face (RF)':
            ws['B22'].value = 'Inlet Size: ' + self.inlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: RF'
        elif self.inlet_facing_2.currentText() == 'Flat Face (FF)':
            ws['B22'].value = 'Inlet Size: ' + self.inlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: FF'
        elif self.inlet_facing_2.currentText() == 'Ring-Type Joint (RTJ)':
            ws['B22'].value = 'Inlet Size: ' + self.inlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: RTJ'
        elif self.inlet_facing_2.currentText() == 'Tongue-and-Groove (T&G)':
            ws['B22'].value = 'Inlet Size: ' + self.inlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: T&G'
        elif self.inlet_facing_2.currentText() == 'Male-and-Female (M&F)':
            ws['B22'].value = 'Inlet Size: ' + self.inlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: M&F'
        elif self.inlet_facing_2.currentText() == 'Other':
            ws['B22'].value = 'Inlet Size: ' + self.inlet_size_2.text() + '     ' + 'Rating: ' + '     ' + self.inlet_facing_3.text()

        if self.outlet_facing_2.currentText() == 'Raised Face (RF)':
            ws['B23'].value = 'Outlet Size: ' + self.outlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: RF'
        elif self.outlet_facing_2.currentText() == 'Flat Face (FF)':
            ws['B23'].value = 'Outlet Size: ' + self.outlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: FF'
        elif self.outlet_facing_2.currentText() == 'Ring-Type Joint (RTJ)':
            ws['B23'].value = 'Outlet Size: ' + self.outlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: RTJ'
        elif self.outlet_facing_2.currentText() == 'Tongue-and-Groove (T&G)':
            ws['B23'].value = 'Outlet Size: ' + self.outlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: T&G'
        elif self.outlet_facing_2.currentText() == 'Male-and-Female (M&F)':
            ws['B23'].value = 'Outlet Size: ' + self.outlet_size_2.text() + '     ' + 'Rating: ' + '     ' + 'Facing: M&F'
        elif self.outlet_facing_2.currentText() == 'Other':
            ws['B23'].value = 'Outlet Size: ' + self.outlet_size_2.text() + '     ' + 'Rating: ' + '     ' + self.outlet_facing_3.text()

        wb.save('Datasheets/Datasheet.xlsx')

    def Materials(self):
        ws['D13'].value = 'Body: ' + self.body_2.text()
        ws['D14'].value = 'Bonnet: ' + self.bonnet_2.text()
        ws['D15'].value = 'Seat (Nozzle): ' + self.seat_2.text() + '                       ' + 'Disk: ' + self.disk_2.text()
        ws['D16'].value = 'Resilient Seat: ' + self.resilient_seat_2.text()
        ws['D17'].value = 'Guide: ' + self.guide_2.text()
        ws['D18'].value = 'Adjusting Ring(s): ' + self.adjusting_rings_2.text()
        ws['D19'].value = 'Spring: ' + self.spring_2.text()
        ws['D20'].value = 'Bellows: ' + self.bellows_2.text()
        ws['D21'].value = 'Balanced Piston: ' + self.balanced_piston_2.text()
        ws['D22'].value = 'Comply with NACE: ' + self.nace_2.currentText()
        ws['D23'].value = 'Internal Gasket: ' + self.internal_gasket_2.text()

        wb.save('Datasheets/Datasheet.xlsx')

    def Accessories(self):
        ws['D27'].value = 'Cap: ' + self.cap_2.currentText()
        ws['D28'].value = 'Lifting Lever: ' + self.lifting_lever_2.currentText()
        ws['D29'].value = 'Test Gag: ' + self.test_gag_2.currentText()
        ws['D30'].value = 'Bug Screen: ' + self.bug_screen_2.currentText()
        ws['D31'].value = 'Other: ' + self.bug_screen_3.text()

        wb.save('Datasheets/Datasheet.xlsx')

    def Service_Conditions(self):
        if self.component_2_name.text() == '':
            ws['B27'].value = 'Fluid: ' + self.component_1_name.text()
        elif self.component_2_name.text() != '':
            ws['B27'].value = 'Fluid: Mixed'

        ws['B28'].value = 'Fluid Phase: ' + self.fluid_phase_2.currentText()
        ws['B29'].value = 'Mass Flow (kg/h): ' + self.mass_flow_2.text()
        ws['B30'].value = 'Relieving Temperature (°C): ' + self.relieving_temp_2.text()
        ws['B31'].value = 'Set Pressure (barg): ' + self.set_pressure_2.text()
        ws['B32'].value = 'Allowable Overpressure: ' + self.overpressure_percentage_2.text() + '%'
        ws['B33'].value = 'Total Back Pressure: ' + self.total_back_pressure_2.text()
        ws['B34'].value = 'Ratio of Specific Heats: ' + self.k_2.text()
        ws['B35'].value = 'Compressibility Factor Z: ' + self.compressibility_factor_2.text()
        ws['B36'].value = self.molecular_weight.text() + ': ' + self.molecular_weight_2.text()
        if self.unit_basis_3.currentText() == 'Mass':
            ws['B37'].value = 'Density at Relieving Pressure (kg/m3): ' + self.density_2.text()
        elif self.unit_basis_3.currentText() == 'Molar':
            ws['B37'].value = 'Density at Relieving Conditions (mol/m3): ' + self.density_2.text()
        ws['B38'].value = 'Viscosity at Relieving Conditions (cP): ' + self.viscosity_2.text()

        wb.save('Datasheets/Datasheet.xlsx')

    def Sizing_And_Selection(self):
        ws['D36'].value = 'Caclulated Orifice Area (mm2): ' + self.discharge_area_2.text()
        ws['D37'].value = 'Selected Effective Orifice Area (mm2): '
        ws['D38'].value = 'Orifice Designation (letter): '
        ws['D39'].value = 'Manufacturer: *'
        ws['D40'].value = 'Model Number: *'
        ws['D41'].value = 'Vendor Calculation Required: Yes'

        wb.save('Datasheets/Datasheet.xlsx')

    def Save_Datasheet(self):
        filename = QFileDialog.getSaveFileName(None, 'Save Datasheet', 'PRV Datasheet.xlsx', 'Excel (*.xlsx)')
        if filename:
            name = filename[0]
            wb.save(name)



# End of code for pressing 'generate datasheet' button ------------------------------------------------------------------------------------------------


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
